import requests

url = "https://liteLLM.ai-coe-test.aws.evernorthcloud.com/v1/chat/completions"

headers = {
    "Authorization": "Bearer YOUR_API_KEY",
    "Content-Type": "application/json"
}

data = {
    "model": "claude-sonnet-4-6",  # confirm exact name from your system
    "messages": [
        {"role": "user", "content": prompt}
    ],
    "temperature": 0.2
}

response = requests.post(url, headers=headers, json=data)

print(response.json())





# =============================================================================
# Teradata Query Usage Metrics Pipeline — Pure Python (No PySpark)
# Works on any Windows VDI / local machine
# =============================================================================
# Dependencies:  pip install pandas sqlglot openpyxl
#
# Output schema:
#   Row_Wid           – incremental integer (1, 2, 3 …)
#   Log_Date          – date of the query  (Metric_Date from CSV)
#   Table_Name        – table referenced in the SQL
#   Column_Name       – column referenced in the SQL
#   Usage_Count       – # times this (table, col) pair appears on that date
#   Distinct_Users    – # unique human users on that date / table / column
#   Distinct_Apps     – # unique app accounts on that date / table / column
#   Created_Timestamp – timestamp when this script ran
#
# App detection rule (case-insensitive prefix match):
#   username starts with  svp | ovt | dt  →  APP
#   everything else                        →  USER
# =============================================================================

import re
from datetime import datetime

import pandas as pd
import sqlglot
from sqlglot import exp
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook

# ── CONFIG ────────────────────────────────────────────────────────────────────

INPUT_CSV   = "input_3_c.csv"          # ← path to your raw CSV from Teradata
OUTPUT_XLSX = "teradata_usage_metrics.xlsx"

APP_PREFIXES = ("svp", "ovt", "dt")   # case-insensitive prefix → app account

# ─────────────────────────────────────────────────────────────────────────────
# STEP 1 — Read & parse the multiline CSV
# Each record looks like:
#   "SELECT ...
#    ...
#    WHERE col = 'x';","YYYY-MM-DD",USERID
# ─────────────────────────────────────────────────────────────────────────────

with open(INPUT_CSV, "r", encoding="utf-8") as fh:
    raw_text = fh.read()

# Drop the header line  (SqlTextInfo,Metric_Date,users)
body = "\n".join(raw_text.splitlines()[1:])

# Regex: capture  (sql_body, date, user)  for every record
RECORD_RE = re.compile(
    r'"(.*?);"\s*,\s*"([^"]+)"\s*,\s*(\S+)',
    re.DOTALL
)
raw_records = RECORD_RE.findall(body)
print(f"[INFO] Records found in CSV : {len(raw_records)}")


# ─────────────────────────────────────────────────────────────────────────────
# STEP 2 — Helpers
# ─────────────────────────────────────────────────────────────────────────────

def classify(username: str) -> str:
    """'app' if username starts with a known app prefix, else 'user'."""
    return "app" if username.lower().startswith(APP_PREFIXES) else "user"


def extract_table_column_pairs(raw_sql: str) -> list[tuple[str, str]]:
    """
    Parse one SQL string and return deduplicated (TABLE, COLUMN) pairs.
    Steps:
      1. Unescape CSV double-quotes  ("" → ")
      2. Strip line comments         (-- …)
      3. Parse AST with sqlglot
      4. Cross-join every table × every column found in the statement
    """
    sql = raw_sql.replace('""', '"').strip()
    sql = re.sub(r'--[^\n]*', '', sql)      # remove  -- comments

    pairs = []
    try:
        statements = sqlglot.parse(sql, error_level=sqlglot.ErrorLevel.IGNORE)
        for stmt in statements:
            if stmt is None:
                continue
            tables = [t.name.upper() for t in stmt.find_all(exp.Table) if t.name]
            cols   = list(dict.fromkeys(     # dedupe, keep first-seen order
                c.name.upper()
                for c in stmt.find_all(exp.Column)
                if c.name
            ))
            for tbl in tables:
                for col in cols:
                    pairs.append((tbl, col))
    except Exception as exc:
        print(f"[WARN] sqlglot error: {exc}")

    return pairs


# ─────────────────────────────────────────────────────────────────────────────
# STEP 3 — Explode every record into (date, table, column, user, acct_type)
# ─────────────────────────────────────────────────────────────────────────────

exploded = []

for raw_sql, date, user in raw_records:
    date = date.strip()
    user = user.strip()
    acct = classify(user)

    pairs = extract_table_column_pairs(raw_sql)
    if not pairs:
        print(f"[WARN] No pairs extracted → date={date}  user={user}")
        continue

    for tbl, col in pairs:
        exploded.append({
            "Log_Date"    : date,
            "Table_Name"  : tbl,
            "Column_Name" : col,
            "username"    : user,
            "acct_type"   : acct,          # "user" | "app"
        })

df_exp = pd.DataFrame(exploded)
print(f"[INFO] Exploded rows (before agg): {len(df_exp)}")


# ─────────────────────────────────────────────────────────────────────────────
# STEP 4 — Aggregate by (Log_Date, Table_Name, Column_Name)
# ─────────────────────────────────────────────────────────────────────────────

grp = df_exp.groupby(["Log_Date", "Table_Name", "Column_Name"], sort=True)

df_agg = grp.agg(
    Usage_Count    = ("username", "count"),
    Distinct_Users = ("username", lambda s:
        df_exp.loc[s.index].loc[df_exp.loc[s.index, "acct_type"] == "user", "username"].nunique()
    ),
    Distinct_Apps  = ("username", lambda s:
        df_exp.loc[s.index].loc[df_exp.loc[s.index, "acct_type"] == "app",  "username"].nunique()
    ),
).reset_index()


# ─────────────────────────────────────────────────────────────────────────────
# STEP 5 — Add Row_Wid and Created_Timestamp
# ─────────────────────────────────────────────────────────────────────────────

df_agg.insert(0, "Row_Wid", range(1, len(df_agg) + 1))
df_agg["Created_Timestamp"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

# Final column order
df_final = df_agg[[
    "Row_Wid",
    "Log_Date",
    "Table_Name",
    "Column_Name",
    "Usage_Count",
    "Distinct_Users",
    "Distinct_Apps",
    "Created_Timestamp",
]]

print(f"[INFO] Final rows: {len(df_final)}")
print(df_final.to_string(index=False))


# ─────────────────────────────────────────────────────────────────────────────
# STEP 6 — Write to Excel with clean formatting
# ─────────────────────────────────────────────────────────────────────────────

df_final.to_excel(OUTPUT_XLSX, index=False, sheet_name="Usage_Metrics")

wb = load_workbook(OUTPUT_XLSX)
ws = wb["Usage_Metrics"]

# Styles
HEADER_FILL   = PatternFill("solid", fgColor="1F4E79")   # dark navy
HEADER_FONT   = Font(name="Arial", bold=True, color="FFFFFF", size=10)
DATA_FONT     = Font(name="Arial", size=10)
ALIGN_CENTER  = Alignment(horizontal="center", vertical="center")
ALIGN_LEFT    = Alignment(horizontal="left",   vertical="center")
THIN_BORDER   = Border(
    left   = Side(style="thin", color="D9D9D9"),
    right  = Side(style="thin", color="D9D9D9"),
    top    = Side(style="thin", color="D9D9D9"),
    bottom = Side(style="thin", color="D9D9D9"),
)
ALT_FILL = PatternFill("solid", fgColor="EBF3FB")        # light-blue alternating rows

# Header row
for cell in ws[1]:
    cell.font      = HEADER_FONT
    cell.fill      = HEADER_FILL
    cell.alignment = ALIGN_CENTER
    cell.border    = THIN_BORDER

ws.row_dimensions[1].height = 22

# Data rows
right_align_cols = {1, 5, 6, 7}    # Row_Wid, Usage_Count, Distinct_*, (numeric)
for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
    fill = ALT_FILL if row_idx % 2 == 0 else PatternFill()
    for cell in row:
        cell.font      = DATA_FONT
        cell.fill      = fill
        cell.border    = THIN_BORDER
        cell.alignment = ALIGN_CENTER if cell.column in right_align_cols else ALIGN_LEFT

# Column widths
col_widths = {
    1: 10,   # Row_Wid
    2: 14,   # Log_Date
    3: 40,   # Table_Name
    4: 35,   # Column_Name
    5: 14,   # Usage_Count
    6: 16,   # Distinct_Users
    7: 14,   # Distinct_Apps
    8: 22,   # Created_Timestamp
}
for col_num, width in col_widths.items():
    ws.column_dimensions[get_column_letter(col_num)].width = width

# Freeze header row
ws.freeze_panes = "A2"

# Auto-filter
ws.auto_filter.ref = ws.dimensions

wb.save(OUTPUT_XLSX)
print(f"\n[INFO] Saved → {OUTPUT_XLSX}")
