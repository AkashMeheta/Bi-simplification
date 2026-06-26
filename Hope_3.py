# =============================================================================
# Teradata Query Usage Metrics Pipeline — Pure Python (No PySpark)
# =============================================================================
# Dependencies:  pip install pandas sqlglot openpyxl
#
# Output schema:
#   Row_Wid           – incremental integer (1, 2, 3 …)
#   Log_Date          – date of the query (Metric_Date from CSV)
#   Table_Name        – table referenced in the SQL
#   Column_Name       – column referenced in the SQL  (* = SELECT * query)
#   Usage_Count       – # times this (table, col) pair appears on that date
#   Distinct_Users    – # unique human users on that date / table / column
#   Distinct_Apps     – # unique app accounts on that date / table / column
#   Created_Timestamp – timestamp when this script ran
#
# App detection rule (case-insensitive prefix match):
#   username starts with  svp | ovt | dt  →  APP,  else  →  USER
# =============================================================================

import re
from datetime import datetime

import pandas as pd
import sqlglot
from sqlglot import exp
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── CONFIG ────────────────────────────────────────────────────────────────────

INPUT_CSV   = "input_3_c.csv"               # ← your raw CSV from Teradata
OUTPUT_XLSX = "teradata_usage_metrics.xlsx"
SKIP_LOG    = "skipped_queries.txt"         # unparseable queries written here

APP_PREFIXES = ("svp", "ovt", "dt")        # case-insensitive prefix → app


# =============================================================================
# STEP 0 — Auto-detect & convert new CSV format to expected input format
# =============================================================================
# New CSV format columns: user_name, Db_nm, Tbl_nm, SqlTextInfo, LogDate,
#                         StartTime, LastResponseTime
#
# Expected input format:  SqlTextInfo,Metric_Date,users
#   Record layout: "<multiline SQL>;","YYYY-MM-DD",USERID
#
# Detection: peek at header — if it contains "user_name" + "logdate" it is
#            the new schema; convert in-memory before STEP 1 runs.
#
# SQL preservation rules:
#   • All internal newlines, commas, and quotes inside the SQL are kept.
#   • SQL is guaranteed to end with exactly one ";" (required by RECORD_RE).
#   • Temp file auto-deleted on process exit via atexit.
# =============================================================================

import csv
import tempfile
import atexit


def _is_new_format(csv_path: str) -> bool:
    """Return True when the CSV uses the new multi-column schema."""
    with open(csv_path, "r", encoding="utf-8") as _fh:
        first_line = _fh.readline()
    header_clean = first_line.strip().lower().replace('"', '')
    return "user_name" in header_clean and "logdate" in header_clean


def _convert_new_csv_to_input_format(src_path: str) -> str:
    """
    Parse the new-format CSV with csv.DictReader (handles quoted multiline
    cells natively) and emit a temp file in the format STEP 1 expects:

        SqlTextInfo,Metric_Date,users
        "<sql ending with ;>","YYYY-MM-DD",USERID
        ...

    All internal newlines, commas, and quote characters inside SqlTextInfo
    are preserved exactly.  Double-quote escaping ("") is applied so that
    the resulting field is a valid RFC-4180 quoted string.
    """
    rows_out = []

    with open(src_path, "r", encoding="utf-8", newline="") as fh:
        reader = csv.DictReader(fh)
        for row in reader:
            sql_raw  = row.get("SqlTextInfo", "").strip()
            log_date = row.get("LogDate", "").strip()
            user     = row.get("user_name", "").strip()

            if not sql_raw or not log_date or not user:
                continue          # skip incomplete / header-only rows

            # Guarantee exactly one trailing semicolon
            sql_norm = sql_raw.rstrip()
            if not sql_norm.endswith(";"):
                sql_norm += ";"

            # Escape internal double-quotes (CSV: " → "")
            sql_esc = sql_norm.replace('"', '""')

            # Assemble record exactly as RECORD_RE expects
            rows_out.append(f'"{sql_esc}","{log_date}",{user}')

    tmp = tempfile.NamedTemporaryFile(
        mode="w", suffix=".csv", delete=False, encoding="utf-8"
    )
    tmp.write("SqlTextInfo,Metric_Date,users\n")
    for rec in rows_out:
        tmp.write(rec + "\n")
    tmp.close()

    atexit.register(os.remove, tmp.name)
    print(f"[INFO] New CSV detected — converted {len(rows_out)} rows → {tmp.name}")
    return tmp.name


# ── Auto-convert (must run BEFORE STEP 1 opens INPUT_CSV) ────────────────────
if _is_new_format(INPUT_CSV):
    INPUT_CSV = _convert_new_csv_to_input_format(INPUT_CSV)
else:
    print("[INFO] Standard CSV format detected — no conversion needed.")


# =============================================================================
# STEP 1 — Read & parse the multiline CSV
# =============================================================================
# Format per record:
#   "SELECT ...          ← opening quote of the SQL field
#    multiline SQL
#    WHERE x = 'Y';","YYYY-MM-DD",USERID   ← record ends here, on ONE line
#
# BUG FIX (regex):
#   Old regex  r'"(.*?);"\s*,\s*"([^"]+)"\s*,\s*(\S+)'  used a non-greedy .*?
#   which stops at the FIRST  ;"  in the file.  If a SQL string literal contains
#   a semicolon (e.g. WHERE code = 'ESI;X') the regex would cut mid-SQL and
#   create thousands of phantom records — explaining the 82 k explosion from
#   only 1 000 real queries.
#
#   Fix: anchor the date field to exactly YYYY-MM-DD format and require the
#   user token to be followed by only whitespace/newline (not more SQL).
#   This ensures we only match real record-end boundaries.
# =============================================================================

with open(INPUT_CSV, "r", encoding="utf-8") as fh:
    raw_text = fh.read()

body = "\n".join(raw_text.splitlines()[1:])     # drop header line

RECORD_RE = re.compile(
    r'"(.*?);"\s*,\s*"(\d{4}-\d{2}-\d{2})"\s*,\s*([^\s,"\n]+)\s*(?:\n|$)',
    #   ^^^^^^^ SQL    ^^^^^^^^^^^^^^^^^^^ date     ^^^^^^^^^^^^^ user (no comma/quote/newline)
    re.DOTALL
)

raw_records = RECORD_RE.findall(body)
print(f"[INFO] Records found in CSV : {len(raw_records)}")


# =============================================================================
# STEP 2 — Helpers
# =============================================================================

def classify(username: str) -> str:
    return "app" if username.lower().startswith(APP_PREFIXES) else "user"


def extract_table_column_pairs(raw_sql: str) -> tuple:
    """
    Returns (pairs, reason):
      pairs  — list of (TABLE, COLUMN) tuples, uppercased
      reason — None on success, string description if the SQL was unparseable

    Three cases handled:
      CASE 1 — Teradata utility commands (SHOW/HELP/COLLECT/EXEC):
               sqlglot returns a Command node with no AST children.
               Genuine metadata ops with nothing to extract → logged & skipped.
      CASE 2 — SELECT *:
               sqlglot returns Star nodes, not Column nodes → record as (TABLE, '*').
      CASE 3 — Normal SELECT/INSERT/UPDATE with explicit columns → standard extraction.
    """
    sql    = raw_sql.replace('""', '"').strip()
    sql    = re.sub(r'--[^\n]*', '', sql)       # strip  -- comments
    pairs  = []
    reason = None

    try:
        statements = sqlglot.parse(
            sql,
            dialect    = "teradata",
            error_level = sqlglot.ErrorLevel.IGNORE
        )
        for stmt in statements:
            if stmt is None:
                continue

            # CASE 1: Teradata utility command
            if type(stmt).__name__ == "Command":
                reason = f"Teradata utility command: {sql.strip()[:80]}"
                continue

            tables = [t.name.upper() for t in stmt.find_all(exp.Table) if t.name]
            cols   = list(dict.fromkeys(
                c.name.upper() for c in stmt.find_all(exp.Column) if c.name
            ))
            stars  = list(stmt.find_all(exp.Star))

            # CASE 2: SELECT *
            if stars and not cols:
                for tbl in tables:
                    pairs.append((tbl, "*"))
                continue

            # CASE 3: Normal query
            for tbl in tables:
                for col in cols:
                    pairs.append((tbl, col))

    except Exception as exc:
        reason = f"Parse error: {exc}"

    return pairs, reason


# =============================================================================
# STEP 3 — Explode every record → flat rows
# =============================================================================

exploded = []
skip_log = []

for raw_sql, date, user in raw_records:
    date = date.strip()
    user = user.strip()
    acct = classify(user)

    pairs, reason = extract_table_column_pairs(raw_sql)

    if reason:
        skip_log.append({"date": date, "user": user, "reason": reason,
                         "sql": raw_sql.replace('""', '"').strip()[:300]})

    for tbl, col in pairs:
        exploded.append({
            "Log_Date"   : date,
            "Table_Name" : tbl,
            "Column_Name": col,
            "username"   : user,
            "acct_type"  : acct,
        })

print(f"[INFO] Exploded rows (before agg) : {len(exploded)}")
print(f"[INFO] Skipped / unparseable      : {len(skip_log)}")

if not exploded:
    raise SystemExit("[ERROR] No rows after parsing — check INPUT_CSV path and format.")

df_exp = pd.DataFrame(exploded)


# =============================================================================
# STEP 4 — Aggregate by (Log_Date, Table_Name, Column_Name)
# =============================================================================
# BUG FIX (aggregation):
#   Old code used a lambda inside groupby.agg() that referenced the outer
#   df_exp DataFrame via closure.  This works on small test data but fails
#   silently on real data because pandas re-indexes groups internally —
#   s.index inside the lambda no longer reliably maps back to df_exp rows,
#   causing every group to see only 1 unique user regardless of actual data.
#
#   Fix: run three SEPARATE groupby passes (total count, users only, apps only)
#   then join the results.  Each pass is clean and index-safe.
# =============================================================================

GRP_KEYS = ["Log_Date", "Table_Name", "Column_Name"]

df_users = (df_exp[df_exp["acct_type"] == "user"]
            .groupby(GRP_KEYS)["username"]
            .nunique()
            .rename("Distinct_Users"))

df_apps  = (df_exp[df_exp["acct_type"] == "app"]
            .groupby(GRP_KEYS)["username"]
            .nunique()
            .rename("Distinct_Apps"))

df_count = (df_exp.groupby(GRP_KEYS)["username"]
            .count()
            .rename("Usage_Count"))

df_agg = (pd.concat([df_count, df_users, df_apps], axis=1)
            .fillna(0)
            .astype({"Distinct_Users": int, "Distinct_Apps": int})
            .reset_index()
            .sort_values(GRP_KEYS)
            .reset_index(drop=True))


# =============================================================================
# STEP 5 — Add Row_Wid and Created_Timestamp
# =============================================================================

df_agg.insert(0, "Row_Wid", range(1, len(df_agg) + 1))
df_agg["Created_Timestamp"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

df_final = df_agg[[
    "Row_Wid", "Log_Date", "Table_Name", "Column_Name",
    "Usage_Count", "Distinct_Users", "Distinct_Apps", "Created_Timestamp",
]]

print(f"[INFO] Final aggregated rows      : {len(df_final)}")
print(df_final.head(10).to_string(index=False))


# =============================================================================
# STEP 6 — Write skip log
# =============================================================================

if skip_log:
    with open(SKIP_LOG, "w", encoding="utf-8") as f:
        f.write(f"Skipped queries — {datetime.now()}\n{'='*80}\n\n")
        for i, entry in enumerate(skip_log, 1):
            f.write(f"[{i}] date={entry['date']}  user={entry['user']}\n")
            f.write(f"     reason : {entry['reason']}\n")
            f.write(f"     sql    : {entry['sql']}\n\n")
    print(f"[INFO] Skip log written → {SKIP_LOG}")


# =============================================================================
# STEP 7 — Write to Excel
# =============================================================================

df_final.to_excel(OUTPUT_XLSX, index=False, sheet_name="Usage_Metrics")

wb = load_workbook(OUTPUT_XLSX)
ws = wb["Usage_Metrics"]

HEADER_FILL  = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT  = Font(name="Arial", bold=True, color="FFFFFF", size=10)
DATA_FONT    = Font(name="Arial", size=10)
ALT_FILL     = PatternFill("solid", fgColor="EBF3FB")
THIN_BORDER  = Border(
    left   = Side(style="thin", color="D9D9D9"),
    right  = Side(style="thin", color="D9D9D9"),
    top    = Side(style="thin", color="D9D9D9"),
    bottom = Side(style="thin", color="D9D9D9"),
)
NUMERIC_COLS = {1, 5, 6, 7}

for cell in ws[1]:
    cell.font      = HEADER_FONT
    cell.fill      = HEADER_FILL
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border    = THIN_BORDER
ws.row_dimensions[1].height = 22

for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
    fill = ALT_FILL if row_idx % 2 == 0 else PatternFill()
    for cell in row:
        cell.font      = DATA_FONT
        cell.fill      = fill
        cell.border    = THIN_BORDER
        cell.alignment = Alignment(
            horizontal = "center" if cell.column in NUMERIC_COLS else "left",
            vertical   = "center"
        )

for col_num, width in {1:10, 2:14, 3:42, 4:36, 5:14, 6:16, 7:14, 8:22}.items():
    ws.column_dimensions[get_column_letter(col_num)].width = width

ws.freeze_panes    = "A2"
ws.auto_filter.ref = ws.dimensions

wb.save(OUTPUT_XLSX)
print(f"[INFO] Saved → {OUTPUT_XLSX}")
