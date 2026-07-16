DEFAULT_TABLE = "SUBQUERIES_CTES"   # placeholder for columns we can't confidently tie to a real base table


def _extract_from_node(node) -> list:
    """
    Pull (table, column) pairs out of a parsed sqlglot node/subtree, using
    scope-based alias resolution instead of a flat find_all() cross-product.

    Anything that can't be confidently tied to one real base table is
    mapped to DEFAULT_TABLE ("SUBQUERIES_CTES") instead of being force-
    attributed to a wrong table or silently dropped. Three cases land here:

      1. A QUALIFIED column whose qualifier resolves to a subquery/CTE
         alias rather than a real exp.Table (e.g. `cte1.col` where cte1 is
         a WITH-clause CTE, not a base table) -- alias_to_table only ever
         contains real exp.Table sources, so this naturally falls through.
      2. An UNQUALIFIED column that's genuinely ambiguous because this
         scope joins more than one real table (e.g. bare `SVC_BEG_DT`
         inside EXTRACT(...) when the FROM has 4 joined tables).
      3. An UNQUALIFIED column whose name matches one of this scope's own
         SELECT-list output aliases (e.g. WHERE DRUG_TYPE_IND <> '?' where
         DRUG_TYPE_IND is a `CASE...END AS DRUG_TYPE_IND` in the SELECT) --
         it's a self-reference to a computed value, not a real table
         column, so it can never be a real (table, column) pair either.

    Everything that CAN be resolved still resolves exactly as before:
      - Each SELECT/JOIN/subquery/CTE is its own scope; columns never leak
        across scopes.
      - Qualified columns (t.col) resolve via THIS scope's own FROM/JOIN
        alias map to a real table when one exists.
      - A single-table scope still attributes bare columns to that table.
    """
    pairs = []

    try:
        root = build_scope(node)
    except Exception:
        root = None

    if root is None:
        tables = [t.name.upper() for t in node.find_all(exp.Table) if t.name]
        cols   = list(dict.fromkeys(c.name.upper() for c in node.find_all(exp.Column) if c.name))
        stars  = list(node.find_all(exp.Star))
        if stars and not cols:
            for tbl in tables:
                pairs.append((tbl, "*"))
        else:
            for tbl in tables:
                for col in cols:
                    pairs.append((tbl, col))
        return pairs

    for scope in root.traverse():
        alias_to_table = {
            alias.upper(): source.name.upper()
            for alias, source in scope.sources.items()
            if isinstance(source, exp.Table) and source.name
        }

        # this scope's own alias set (real tables + subquery/CTE aliases)
        # -- used only to know whether a qualifier belongs to THIS scope
        # at all vs. being a completely stray/unrecognized reference.
        all_scope_aliases = {alias.upper() for alias in scope.sources.keys()}

        scope_tables = list(dict.fromkeys(alias_to_table.values()))

        output_alias_names = {
            sel.alias.upper() for sel in getattr(scope.expression, "selects", [])
            if isinstance(sel, exp.Alias)
        }

        has_star = any(
            isinstance(sel, exp.Star) or
            (isinstance(sel, exp.Column) and isinstance(sel.this, exp.Star))
            for sel in getattr(scope.expression, "selects", [])
        )

        scope_columns = [c for c in scope.columns if c.name]

        if has_star and not scope_columns:
            for tbl in scope_tables:
                pairs.append((tbl, "*"))
            continue

        seen = set()
        for col in scope_columns:
            col_name = col.name.upper()
            tbl_ref  = col.table.upper() if col.table else None

            if tbl_ref:
                real_tbl = alias_to_table.get(tbl_ref)
                if real_tbl:
                    key = (real_tbl, col_name)
                else:
                    # qualifier exists but isn't a real base table in this
                    # scope -> it's a subquery/CTE alias
                    key = (DEFAULT_TABLE, col_name)
                if key not in seen:
                    seen.add(key)
                    pairs.append(key)
                continue

            # unqualified column
            if col_name in output_alias_names:
                key = (DEFAULT_TABLE, col_name)
            elif len(scope_tables) == 1:
                key = (scope_tables[0], col_name)
            else:
                key = (DEFAULT_TABLE, col_name)

            if key not in seen:
                seen.add(key)
                pairs.append(key)

    return pairs

=============================================================================
# Teradata Query Usage Metrics Pipeline — Pure Python (No PySpark)
# =============================================================================
# Dependencies:  pip install pandas sqlglot openpyxl
#
# Output schema:
#   Row_Wid           – incremental integer (1, 2, 3 …)
#   Log_Date          – date of the query (Metric_Date from CSV)
#   Table_Name        – table referenced in the SQL
#   Column_Name       – column referenced in the SQL  (* = SELECT * query)
#   Usage_Count       – # times this (table, col) pair appears on that date
#   Distinct_Users    – # unique human users on that date / table / column
#   Distinct_Apps     – # unique app accounts on that date / table / column
#   Created_Timestamp – timestamp when this script ran
#
# App detection rule (case-insensitive prefix match):
#   username starts with  svp | ovt | dt  →  APP,  else  →  USER
# =============================================================================

import re
import logging
from datetime import datetime

import pandas as pd
import sqlglot
from sqlglot import exp
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Silence sqlglot's benign "falling back to Command" warnings — we handle
# that fallback ourselves in extract_table_column_pairs() below, so the
# raw warning text would just be noise (one line per recovered query).
logging.getLogger("sqlglot").setLevel(logging.ERROR)

# ── CONFIG ────────────────────────────────────────────────────────────────────

INPUT_CSV   = "input_3_c.csv"               # ← your raw CSV from Teradata
OUTPUT_XLSX = "teradata_usage_metrics.xlsx"
SKIP_LOG    = "skipped_queries.txt"         # unparseable queries written here

APP_PREFIXES = ("svp", "ovt", "dt")        # case-insensitive prefix → app


# =============================================================================
# STEP 0 — Auto-detect & convert alternate CSV formats to expected input format
# =============================================================================
# Supports multiple input schemas, normalized into what STEP 1 expects:
#
#       SqlTextInfo,Metric_Date,users
#       "<multiline SQL ending in ;>","YYYY-MM-DD",USERID
#
# Format A — "new_format":
#   Columns: user_name, Db_nm, Tbl_nm, SqlTextInfo, LogDate, StartTime,
#            LastResponseTime
#   Detection: header contains "user_name" AND "logdate"
#
# Format B/C — "sql_users_format":
#   Columns: SqlTextInfo, Metric_Date, users — handled regardless of quoting:
#   fully quoted (standard), fully unquoted (select * from t1,29/02/2026,user1),
#   or a MIX of both within the same file (a CSV writer using RFC-4180 minimal
#   quoting only wraps a field in quotes when it needs to, e.g. the SQL
#   contains a comma — so some rows may be quoted and others not). Every row
#   is inspected and normalized on its own via csv.reader, so mixed quoting
#   is handled correctly instead of guessing the whole file's format from
#   just the first row.
#
# Dates are normalized to YYYY-MM-DD (required by STEP 1's RECORD_RE),
# handling DD/MM/YYYY and ISO-8601 timestamps like 2026-06-25T14:48:39.466Z.
#
# csv.field_size_limit() is also raised, since the default (~131072 bytes)
# is too small for large multiline SQL text fields in big exports.
# =============================================================================

import csv
import os
import sys
import tempfile
import atexit

# Default csv field size limit (~131072 bytes) is too small for large
# multiline SQL text fields in big exports — raise it. sys.maxsize can raise
# OverflowError on platforms where the C `long` is 32-bit (e.g. some Windows
# builds), so step down until it's accepted.
_new_limit = sys.maxsize
while True:
    try:
        csv.field_size_limit(_new_limit)
        break
    except OverflowError:
        _new_limit = int(_new_limit / 10)


def _detect_source_encoding(path: str, sample_size: int = 1_000_000) -> str:
    """
    Detect a workable text encoding for the SOURCE CSV (the raw export,
    before we normalize it into our own always-UTF-8 temp file).

    Enterprise CSV exports (Teradata Studio, SQL Assistant, Excel,
    PowerShell `Export-Csv` on Windows) are very often Windows-1252
    (cp1252) rather than UTF-8. That's what produces errors like:

        UnicodeDecodeError: 'utf-8' codec can't decode byte 0x97 ...

    0x97 is an em-dash / smart-quote character under cp1252 — common in
    pasted comments or copy/pasted query text — but is not valid UTF-8 on
    its own.

    We sample the file (not the whole thing — large files could be 100s
    of MB) and try candidate encodings in order of how likely/safe they
    are. latin-1 (ISO-8859-1) is listed last as a guaranteed fallback:
    it maps every byte 0-255 to a code point, so it can never raise
    UnicodeDecodeError, even if it's not the "true" encoding.
    """
    candidates = ("utf-8-sig", "utf-8", "cp1252", "latin-1")
    with open(path, "rb") as fh:
        sample = fh.read(sample_size)

    for enc in candidates:
        try:
            sample.decode(enc)
            return enc
        except UnicodeDecodeError:
            continue
    return "latin-1"  # unreachable in practice — latin-1 never raises


def _peek_lines(csv_path: str, n: int = 2, encoding: str = "utf-8"):
    lines = []
    with open(csv_path, "r", encoding=encoding, errors="replace") as fh:
        for line in fh:
            if line.strip():
                lines.append(line)
            if len(lines) >= n:
                break
    return lines


def _detect_format(csv_path: str, encoding: str = "utf-8") -> str:
    """
    Only two schemas need distinguishing up front:
      • "new_format"       — user_name/Db_nm/Tbl_nm/SqlTextInfo/LogDate/... columns
      • "sql_users_format" — SqlTextInfo,Metric_Date,users columns (quoted,
                              unquoted, or a MIX of both within the same file)

    We deliberately do NOT try to decide "already fully quoted" vs "needs
    conversion" here. A CSV writer using RFC-4180 minimal quoting only wraps
    a field in quotes when it actually needs to — so a real file can easily
    have some rows quoted and others not. Peeking at just the first data row
    to guess the whole file's format misclassifies that kind of mixed file
    and silently drops every row after the first mismatch. Instead,
    _convert_sql_users_csv_to_input_format() below inspects EVERY row on
    its own and normalizes it individually, so mixed quoting is a non-issue.
    """
    lines = _peek_lines(csv_path, 1, encoding=encoding)
    if not lines:
        return "sql_users_format"

    header_clean = lines[0].strip().lower().replace('"', '')

    if "user_name" in header_clean and "logdate" in header_clean:
        return "new_format"

    return "sql_users_format"


def _normalize_date(date_str: str) -> str:
    """
    Normalize a date string to YYYY-MM-DD, the format STEP 1's RECORD_RE
    requires. Handles:
      • YYYY-MM-DD                          -> unchanged
      • YYYY-MM-DDTHH:MM:SS(.ffffff)?Z?      -> date part only
                                                (2026-06-25T14:48:39.466Z -> 2026-06-25)
      • DD/MM/YYYY                           -> rearranged (29/02/2026 -> 2026-02-29)
      • YYYY/MM/DD                           -> rearranged
    Unrecognized shapes are returned unchanged.
    """
    date_str = date_str.strip()

    if re.match(r'^\d{4}-\d{2}-\d{2}$', date_str):
        return date_str

    m = re.match(r'^(\d{4}-\d{2}-\d{2})[T ]\d{2}:\d{2}:\d{2}', date_str)
    if m:
        return m.group(1)

    m = re.match(r'^(\d{1,2})/(\d{1,2})/(\d{4})$', date_str)
    if m:
        dd, mm, yyyy = m.groups()
        return f"{yyyy}-{mm.zfill(2)}-{dd.zfill(2)}"

    m = re.match(r'^(\d{4})/(\d{1,2})/(\d{1,2})$', date_str)
    if m:
        yyyy, mm, dd = m.groups()
        return f"{yyyy}-{mm.zfill(2)}-{dd.zfill(2)}"

    return date_str


def _write_temp_csv(rows_out, source_label: str) -> str:
    tmp = tempfile.NamedTemporaryFile(
        mode="w", suffix=".csv", delete=False, encoding="utf-8"
    )
    tmp.write("SqlTextInfo,Metric_Date,users\n")
    for rec in rows_out:
        tmp.write(rec + "\n")
    tmp.close()

    atexit.register(os.remove, tmp.name)
    print(f"[INFO] {source_label} detected — converted {len(rows_out)} rows -> {tmp.name}")
    return tmp.name


def _normalize_sql_field(sql_raw: str) -> str:
    """Guarantee exactly one trailing ';' and escape internal double-quotes
    so the field is a valid RFC-4180 quoted string."""
    sql_norm = sql_raw.strip().rstrip()
    if not sql_norm.endswith(";"):
        sql_norm += ";"
    return sql_norm.replace('"', '""')


def _convert_new_csv_to_input_format(src_path: str, encoding: str = "utf-8") -> str:
    rows_out = []
    # errors="replace" is a last-resort safety net: even after picking the
    # best-guess encoding from a sample, a handful of stray bytes further
    # into a 160k-row file could still be off. Rather than crash the whole
    # pipeline on one bad byte, swap it for U+FFFD and keep going.
    with open(src_path, "r", encoding=encoding, errors="replace", newline="") as fh:
        reader = csv.DictReader(fh)
        for row in reader:
            sql_raw  = (row.get("SqlTextInfo") or "").strip()
            log_date = _normalize_date((row.get("LogDate") or "").strip())
            user     = (row.get("user_name") or "").strip()

            if not sql_raw or not log_date or not user:
                continue

            sql_esc = _normalize_sql_field(sql_raw)
            rows_out.append(f'"{sql_esc}","{log_date}",{user}')

    return _write_temp_csv(rows_out, "New CSV schema")


def _convert_sql_users_csv_to_input_format(src_path: str, encoding: str = "utf-8") -> str:
    """
    Handles SqlTextInfo,Metric_Date,users files regardless of quoting —
    fully quoted (standard), fully unquoted, or a MIX of both across rows.

    csv.reader() respects RFC-4180 quoting per field/row (including
    embedded newlines inside quoted multiline SQL), so each row is read
    correctly whether or not that particular row happened to need quotes.

    The only case csv.reader can't disambiguate on its own is a row whose
    SQL contains a comma that was NOT quoted/escaped at all — that row
    parses into MORE than 3 fields. For those, we recombine: the last two
    fields are always Metric_Date and users (neither contains commas), so
    everything else — however many stray commas it has — is rejoined back
    into the SQL text.
    """
    rows_out = []
    with open(src_path, "r", encoding=encoding, errors="replace", newline="") as fh:
        reader = csv.reader(fh)
        header_skipped = False
        for fields in reader:
            if not fields or not any((f or "").strip() for f in fields):
                continue  # skip blank lines

            if not header_skipped:
                header_skipped = True
                continue  # skip header row

            if len(fields) == 3:
                sql_raw, log_date, user = fields
            elif len(fields) > 3:
                # Unescaped comma(s) inside an unquoted SQL field — rejoin.
                sql_raw  = ",".join(fields[:-2])
                log_date = fields[-2]
                user     = fields[-1]
            else:
                continue  # malformed row — not enough fields, skip

            sql_raw  = sql_raw.strip()
            log_date = _normalize_date(log_date.strip())
            user     = user.strip()

            if not sql_raw or not log_date or not user:
                continue

            sql_esc = _normalize_sql_field(sql_raw)
            rows_out.append(f'"{sql_esc}","{log_date}",{user}')

    return _write_temp_csv(rows_out, "SqlTextInfo/Metric_Date/users CSV")


_src_encoding = _detect_source_encoding(INPUT_CSV)
print(f"[INFO] Source CSV encoding detected as: {_src_encoding}")

_fmt = _detect_format(INPUT_CSV, encoding=_src_encoding)

if _fmt == "new_format":
    INPUT_CSV = _convert_new_csv_to_input_format(INPUT_CSV, encoding=_src_encoding)
else:
    INPUT_CSV = _convert_sql_users_csv_to_input_format(INPUT_CSV, encoding=_src_encoding)


# =============================================================================
# STEP 1 — Read & parse the multiline CSV
# =============================================================================
# Record format:
#   "SELECT ...
#    multiline SQL ...
#    WHERE x='Y';","YYYY-MM-DD",USERID
#
# NOTE on approach: this file was previously read in full into one giant
# string and matched with a DOTALL regex over the whole body. That doesn't
# scale well to 160k+ rows with very large SqlTextInfo fields — it holds
# the entire file (and every intermediate regex match) in memory at once,
# and a single giant backtracking regex over a multi-hundred-MB string is
# slow and can be fragile.
#
# INPUT_CSV at this point is ALWAYS our own temp file written by
# _write_temp_csv() in STEP 0 — always UTF-8, always valid RFC-4180 CSV
# (quoted SQL field with internal quotes escaped as "", quoted date,
# unquoted user). That means we can just stream it row-by-row with
# csv.reader (which already has field_size_limit raised above to handle
# very large SqlTextInfo values), instead of re-parsing it with regex.
# This keeps memory bounded to one row at a time rather than the whole
# file, and is dramatically faster on large datasets.
# =============================================================================

raw_records = []
_skipped_bad_rows = 0

with open(INPUT_CSV, "r", encoding="utf-8", newline="") as fh:
    reader = csv.reader(fh)
    next(reader, None)  # drop header row written by _write_temp_csv

    for fields in reader:
        if len(fields) < 3:
            _skipped_bad_rows += 1
            continue

        sql_field, date_field, user_field = fields[0], fields[1], fields[2]
        date_field = date_field.strip()

        if not re.match(r'^\d{4}-\d{2}-\d{2}$', date_field):
            _skipped_bad_rows += 1
            continue

        if not sql_field.strip() or not user_field.strip():
            _skipped_bad_rows += 1
            continue

        raw_records.append((sql_field, date_field, user_field))

print(f"[INFO] Records found in CSV : {len(raw_records)}")
if _skipped_bad_rows:
    print(f"[INFO] Rows skipped (malformed) : {_skipped_bad_rows}")

# =============================================================================
# STEP 2 — Helpers
# =============================================================================

def classify(username: str) -> str:
    return "app" if username.lower().startswith(APP_PREFIXES) else "user"


def _strip_locking_prefix(text: str) -> str:
    """
    Teradata locking-request modifiers (e.g. LOCKING ROW FOR ACCESS) are
    valid Teradata syntax at the TOP level of a statement, but sqlglot can't
    parse them when they appear nested inside a CREATE ... AS (...) body —
    that nesting is what causes the whole CREATE to fall back to a generic,
    unparseable "Command" node. Since the clause is just a lock hint (it
    doesn't affect which tables/columns are referenced), we strip it before
    re-parsing the inner query on its own.
    """
    stripped = text.lstrip()
    if re.match(r'(?i)^locking\b', stripped):
        m = re.search(r'(?i)\b(select|with)\b', stripped)
        if m:
            return stripped[m.start():]
    return text


def _extract_inner_query(sql_text: str):
    """
    Pull the SELECT/WITH body out of a CREATE ... AS (...) / CREATE ... AS
    SELECT ... statement so it can be parsed on its own, independent of the
    (possibly-unsupported-by-sqlglot) outer CREATE syntax.
    """
    m = re.search(r'\bAS\b\s*\(', sql_text, flags=re.IGNORECASE)
    if m:
        start = m.end() - 1  # index of '('
        depth = 0
        for i in range(start, len(sql_text)):
            if sql_text[i] == '(':
                depth += 1
            elif sql_text[i] == ')':
                depth -= 1
                if depth == 0:
                    return _strip_locking_prefix(sql_text[start + 1:i].strip())
        return None  # unbalanced parens — give up

    m = re.search(r'\bAS\b\s*(SELECT\b.*)', sql_text, flags=re.IGNORECASE | re.DOTALL)
    if m:
        return _strip_locking_prefix(m.group(1).strip())
    return None


def _extract_from_node(node) -> list:
    """Pull (table, column) pairs out of a parsed sqlglot node/subtree."""
    tables = [t.name.upper() for t in node.find_all(exp.Table) if t.name]
    cols   = list(dict.fromkeys(c.name.upper() for c in node.find_all(exp.Column) if c.name))
    stars  = list(node.find_all(exp.Star))

    pairs = []
    if stars and not cols:
        for tbl in tables:
            pairs.append((tbl, "*"))
    else:
        for tbl in tables:
            for col in cols:
                pairs.append((tbl, col))
    return pairs


def extract_table_column_pairs(raw_sql: str) -> tuple:
    """
    Returns (pairs, reason):
      pairs  — list of (TABLE, COLUMN) tuples, uppercased
      reason — None on success, string description if the SQL was unparseable

    Cases handled:
      CASE 1 — Genuine Teradata utility commands (SHOW/HELP/COLLECT/EXEC):
               sqlglot returns a Command node with no AST children.
               Logged & skipped, same as before.
      CASE 1b — CREATE ... AS (...) wrapping a LOCKING ROW FOR ACCESS (or
               similar) clause: sqlglot also returns Command for the WHOLE
               statement here, even though the inner query is perfectly
               valid Teradata SQL on its own. We recover by extracting the
               inner SELECT/WITH body, stripping the locking clause, and
               re-parsing just that — using ONLY the Teradata dialect (no
               cross-dialect fallback, to avoid other dialects silently
               "recovering" genuinely malformed SQL as something benign).
      CASE 2 — SELECT *: sqlglot returns Star nodes, not Column nodes →
               recorded as (TABLE, '*').
      CASE 3 — Normal SELECT/INSERT/UPDATE, and CREATE ... AS (SELECT ...)
               that parses fine directly → standard extraction, restricted
               to the query subtree (search_root) so the object being
               CREATEd isn't itself counted as a "referenced" table.
    """
    sql    = raw_sql.replace('""', '"').strip()
    sql    = re.sub(r'--[^\n]*', '', sql)       # strip  -- comments
    pairs  = []
    reason = None

    try:
        statements = sqlglot.parse(
            sql,
            dialect     = "teradata",
            error_level = sqlglot.ErrorLevel.IGNORE
        )
        for stmt in statements:
            if stmt is None:
                continue

            # CASE 1 / 1b: Teradata Command fallback
            if type(stmt).__name__ == "Command":
                inner_sql = _extract_inner_query(sql)
                recovered_pairs = []
                if inner_sql:
                    try:
                        inner_statements = sqlglot.parse(
                            inner_sql,
                            dialect     = "teradata",
                            error_level = sqlglot.ErrorLevel.IGNORE
                        )
                    except Exception:
                        inner_statements = []
                    if inner_statements and not any(
                        type(s).__name__ == "Command" for s in inner_statements if s is not None
                    ):
                        for inner_stmt in inner_statements:
                            if inner_stmt is not None:
                                recovered_pairs.extend(_extract_from_node(inner_stmt))

                if recovered_pairs:
                    pairs.extend(recovered_pairs)   # recovered — not a real skip
                else:
                    reason = f"Teradata utility command: {sql.strip()[:80]}"
                continue

            # search_root: for CREATE ... AS (SELECT ...), only look inside
            # the query part so the view/table name being CREATEd isn't
            # itself counted as a "referenced" table.
            search_root = stmt
            if isinstance(stmt, exp.Create):
                inner_expr = stmt.args.get("expression")
                if inner_expr is not None:
                    search_root = inner_expr

            pairs.extend(_extract_from_node(search_root))

    except Exception as exc:
        reason = f"Parse error: {exc}"

    if pairs:
        reason = None

    return pairs, reason


# =============================================================================
# STEP 3 — Explode every record → flat rows
# =============================================================================

exploded = []
skip_log = []
_total_records = len(raw_records)
_progress_every = 20_000  # print a heartbeat every N records on large runs

for _i, (raw_sql, date, user) in enumerate(raw_records, start=1):
    date = date.strip()
    user = user.strip()
    acct = classify(user)

    pairs, reason = extract_table_column_pairs(raw_sql)

    if reason:
        skip_log.append({"date": date, "user": user, "reason": reason,
                         "sql": raw_sql.replace('""', '"').strip()[:300]})

    for tbl, col in pairs:
        exploded.append({
            "Log_Date"   : date,
            "Table_Name" : tbl,
            "Column_Name": col,
            "username"   : user,
            "acct_type"  : acct,
        })

    if _i % _progress_every == 0 or _i == _total_records:
        print(f"[INFO] Parsed {_i:,} / {_total_records:,} records "
              f"({len(exploded):,} rows exploded so far)")

print(f"[INFO] Exploded rows (before agg) : {len(exploded)}")
print(f"[INFO] Skipped / unparseable      : {len(skip_log)}")

if not exploded:
    raise SystemExit("[ERROR] No rows after parsing — check INPUT_CSV path and format.")

df_exp = pd.DataFrame(exploded)


# =============================================================================
# STEP 4 — Aggregate by (Log_Date, Table_Name, Column_Name)
# =============================================================================

GRP_KEYS = ["Log_Date", "Table_Name", "Column_Name"]

df_users = (df_exp[df_exp["acct_type"] == "user"]
            .groupby(GRP_KEYS)["username"]
            .nunique()
            .rename("Distinct_Users"))

df_apps  = (df_exp[df_exp["acct_type"] == "app"]
            .groupby(GRP_KEYS)["username"]
            .nunique()
            .rename("Distinct_Apps"))

df_count = (df_exp.groupby(GRP_KEYS)["username"]
            .count()
            .rename("Usage_Count"))

df_agg = (pd.concat([df_count, df_users, df_apps], axis=1)
            .fillna(0)
            .astype({"Distinct_Users": int, "Distinct_Apps": int})
            .reset_index()
            .sort_values(GRP_KEYS)
            .reset_index(drop=True))


# =============================================================================
# STEP 5 — Add Row_Wid and Created_Timestamp
# =============================================================================

df_agg.insert(0, "Row_Wid", range(1, len(df_agg) + 1))
df_agg["Created_Timestamp"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

df_final = df_agg[[
    "Row_Wid", "Log_Date", "Table_Name", "Column_Name",
    "Usage_Count", "Distinct_Users", "Distinct_Apps", "Created_Timestamp",
]]

print(f"[INFO] Final aggregated rows      : {len(df_final)}")
print(df_final.head(10).to_string(index=False))


# =============================================================================
# STEP 6 — Write skip log
# =============================================================================

if skip_log:
    with open(SKIP_LOG, "w", encoding="utf-8") as f:
        f.write(f"Skipped queries — {datetime.now()}\n{'='*80}\n\n")
        for i, entry in enumerate(skip_log, 1):
            f.write(f"[{i}] date={entry['date']}  user={entry['user']}\n")
            f.write(f"     reason : {entry['reason']}\n")
            f.write(f"     sql    : {entry['sql']}\n\n")
    print(f"[INFO] Skip log written → {SKIP_LOG}")


# =============================================================================
# STEP 7 — Write to Excel
# =============================================================================

df_final.to_excel(OUTPUT_XLSX, index=False, sheet_name="Usage_Metrics")

wb = load_workbook(OUTPUT_XLSX)
ws = wb["Usage_Metrics"]

HEADER_FILL  = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT  = Font(name="Arial", bold=True, color="FFFFFF", size=10)
DATA_FONT    = Font(name="Arial", size=10)
ALT_FILL     = PatternFill("solid", fgColor="EBF3FB")
THIN_BORDER  = Border(
    left   = Side(style="thin", color="D9D9D9"),
    right  = Side(style="thin", color="D9D9D9"),
    top    = Side(style="thin", color="D9D9D9"),
    bottom = Side(style="thin", color="D9D9D9"),
)
NUMERIC_COLS = {1, 5, 6, 7}

for cell in ws[1]:
    cell.font      = HEADER_FONT
    cell.fill      = HEADER_FILL
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border    = THIN_BORDER
ws.row_dimensions[1].height = 22

for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
    fill = ALT_FILL if row_idx % 2 == 0 else PatternFill()
    for cell in row:
        cell.font      = DATA_FONT
        cell.fill      = fill
        cell.border    = THIN_BORDER
        cell.alignment = Alignment(
            horizontal = "center" if cell.column in NUMERIC_COLS else "left",
            vertical   = "center"
        )

for col_num, width in {1:10, 2:14, 3:42, 4:36, 5:14, 6:16, 7:14, 8:22}.items():
    ws.column_dimensions[get_column_letter(col_num)].width = width

ws.freeze_panes    = "A2"
ws.auto_filter.ref = ws.dimensions

wb.save(OUTPUT_XLSX)
print(f"[INFO] Saved → {OUTPUT_XLSX}")



from sqlglot.optimizer.scope import build_scope
def _extract_from_node(node) -> list:
    """
    Pull (table, column) pairs out of a parsed sqlglot node/subtree.

    IMPORTANT: this uses sqlglot's scope resolution (build_scope) instead
    of a flat find_all() cross-product. The old approach collected every
    exp.Table and every exp.Column anywhere in the subtree and paired ALL
    of them together — which is correct for single-table queries but
    silently corrupts multi-table queries (JOINs, subqueries, CTEs) by
    attaching columns to tables they don't belong to. With ~160k real
    queries (lots of joins), that cross-contamination was significant
    enough to distort the usage metrics.

    Scope-based resolution instead:
      - Walks each SELECT scope independently (outer query, each JOINed
        table, each subquery, each CTE body are separate scopes), so
        columns from one scope never get attributed to tables that only
        appear in a different scope.
      - For a QUALIFIED column (t.col), resolves alias `t` to its real
        table name using that scope's own FROM/JOIN sources.
      - For an UNQUALIFIED column, it's still ambiguous which table (of
        possibly several) in that scope it belongs to, so it's attributed
        to every table in THAT scope only — not the whole statement.
      - A qualifier that doesn't resolve to a real table in this scope
        (e.g. it refers to a CTE alias rather than a base table) is
        skipped rather than guessed at.
    """
    pairs = []

    try:
        root = build_scope(node)
    except Exception:
        root = None

    if root is None:
        # Fallback for fragments build_scope can't handle — keep the old
        # best-effort behaviour rather than dropping the row entirely.
        tables = [t.name.upper() for t in node.find_all(exp.Table) if t.name]
        cols   = list(dict.fromkeys(c.name.upper() for c in node.find_all(exp.Column) if c.name))
        stars  = list(node.find_all(exp.Star))
        if stars and not cols:
            for tbl in tables:
                pairs.append((tbl, "*"))
        else:
            for tbl in tables:
                for col in cols:
                    pairs.append((tbl, col))
        return pairs

    for scope in root.traverse():
        # alias -> real table name, restricted to THIS scope's own
        # FROM / JOIN sources (CTE/subquery sources are Scope objects,
        # not exp.Table, and are deliberately excluded here).
        alias_to_table = {
            alias.upper(): source.name.upper()
            for alias, source in scope.sources.items()
            if isinstance(source, exp.Table) and source.name
        }
        if not alias_to_table:
            continue

        scope_tables = list(dict.fromkeys(alias_to_table.values()))

        # scope.columns = columns belonging to THIS scope only (sqlglot
        # excludes columns that live inside nested subquery/CTE scopes).
        scope_columns = [c for c in scope.columns if c.name]

        has_star = any(
            isinstance(sel, exp.Star) or
            (isinstance(sel, exp.Column) and isinstance(sel.this, exp.Star))
            for sel in getattr(scope.expression, "selects", [])
        )

        if has_star and not scope_columns:
            for tbl in scope_tables:
                pairs.append((tbl, "*"))
            continue

        seen = set()
        for col in scope_columns:
            col_name = col.name.upper()
            tbl_ref  = col.table.upper() if col.table else None

            if tbl_ref:
                real_tbl = alias_to_table.get(tbl_ref)
                targets  = [real_tbl] if real_tbl else []
            else:
                # unqualified — ambiguous only within this scope's own
                # tables, never the rest of the statement
                targets = scope_tables

            for tbl in targets:
                key = (tbl, col_name)
                if key not in seen:
                    seen.add(key)
                    pairs.append(key)

    return pairs
