# =============================================================================
# Teradata Query Usage Metrics Pipeline — Pure Python (No PySpark)
# =============================================================================
# Dependencies:  pip install pandas sqlglot openpyxl
#
# Output schema:
#   Row_Wid           – incremental integer (1, 2, 3 …)
#   Log_Date          – date of the query (Metric_Date from CSV)
#   Table_Name        – table referenced in the SQL
#   Column_Name       – column referenced in the SQL  (* = SELECT * query)
#   Usage_Count       – # times this (table, col) pair appears on that date
#   Distinct_Users    – # unique human users on that date / table / column
#   Distinct_Apps     – # unique app accounts on that date / table / column
#   Created_Timestamp – timestamp when this script ran
#
# App detection rule (case-insensitive prefix match):
#   username starts with  svp | ovt | dt  →  APP,  else  →  USER
# =============================================================================

import re
import logging
from datetime import datetime

import pandas as pd
import sqlglot
from sqlglot import exp
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Silence sqlglot's benign "falling back to Command" warnings — we handle
# that fallback ourselves in extract_table_column_pairs() below, so the
# raw warning text would just be noise (one line per recovered query).
logging.getLogger("sqlglot").setLevel(logging.ERROR)

# ── CONFIG ────────────────────────────────────────────────────────────────────

INPUT_CSV   = "input_3_c.csv"               # ← your raw CSV from Teradata
OUTPUT_XLSX = "teradata_usage_metrics.xlsx"
SKIP_LOG    = "skipped_queries.txt"         # unparseable queries written here

APP_PREFIXES = ("svp", "ovt", "dt")        # case-insensitive prefix → app


# =============================================================================
# STEP 0 — Auto-detect & convert alternate CSV formats to expected input format
# =============================================================================
# Supports THREE input schemas, normalized into what STEP 1 expects:
#
#       SqlTextInfo,Metric_Date,users
#       "<multiline SQL ending in ;>","YYYY-MM-DD",USERID
#
# Format A — "new_format":
#   Columns: user_name, Db_nm, Tbl_nm, SqlTextInfo, LogDate, StartTime,
#            LastResponseTime
#   Detection: header contains "user_name" AND "logdate"
#
# Format B — "standard_format" (already correct, passthrough):
#   Columns: SqlTextInfo, Metric_Date, users  (already quoted + ; terminated)
#
# Format C — "unquoted_format":
#   Columns: SqlTextInfo, Metric_Date, users, but UNQUOTED, no semicolon:
#       select * from t1,29/02/2026,user1
#   We rsplit(",", 2) from the right (last two fields are always date/user;
#   everything else, however many commas, is the SQL) and re-quote/escape.
#
# Dates are normalized to YYYY-MM-DD (required by STEP 1's RECORD_RE),
# handling DD/MM/YYYY and ISO-8601 timestamps like 2026-06-25T14:48:39.466Z.
# =============================================================================

import csv
import os
import sys
import tempfile
import atexit

# Default csv field size limit (~131072 bytes) is too small for large
# multiline SQL text fields in big exports — raise it. sys.maxsize can raise
# OverflowError on platforms where the C `long` is 32-bit (e.g. some Windows
# builds), so step down until it's accepted.
_new_limit = sys.maxsize
while True:
    try:
        csv.field_size_limit(_new_limit)
        break
    except OverflowError:
        _new_limit = int(_new_limit / 10)


def _peek_lines(csv_path: str, n: int = 2):
    lines = []
    with open(csv_path, "r", encoding="utf-8") as fh:
        for line in fh:
            if line.strip():
                lines.append(line)
            if len(lines) >= n:
                break
    return lines


def _detect_format(csv_path: str) -> str:
    """
    Only two schemas need distinguishing up front:
      • "new_format"       — user_name/Db_nm/Tbl_nm/SqlTextInfo/LogDate/... columns
      • "sql_users_format" — SqlTextInfo,Metric_Date,users columns (quoted,
                              unquoted, or a MIX of both within the same file)

    We deliberately do NOT try to decide "already fully quoted" vs "needs
    conversion" here anymore. A CSV writer using RFC-4180 minimal quoting
    only wraps a field in quotes when it actually needs to (e.g. the SQL
    text happens to contain a comma) — so a real file can easily have some
    rows quoted and others not. Peeking at just the first data row to guess
    the whole file's format misclassifies that kind of mixed file and
    silently drops every row after the first mismatch. Instead,
    _convert_sql_users_csv_to_input_format() below inspects EVERY row on
    its own and normalizes it individually, so mixed quoting is a non-issue.
    """
    lines = _peek_lines(csv_path, 1)
    if not lines:
        return "sql_users_format"

    header_clean = lines[0].strip().lower().replace('"', '')

    if "user_name" in header_clean and "logdate" in header_clean:
        return "new_format"

    return "sql_users_format"


def _normalize_date(date_str: str) -> str:
    """
    Normalize a date string to YYYY-MM-DD, the format STEP 1's RECORD_RE
    requires. Handles:
      • YYYY-MM-DD                          -> unchanged
      • YYYY-MM-DDTHH:MM:SS(.ffffff)?Z?      -> date part only
                                                (2026-06-25T14:48:39.466Z -> 2026-06-25)
      • DD/MM/YYYY                           -> rearranged (29/02/2026 -> 2026-02-29)
      • YYYY/MM/DD                           -> rearranged
    Unrecognized shapes are returned unchanged.
    """
    date_str = date_str.strip()

    if re.match(r'^\d{4}-\d{2}-\d{2}$', date_str):
        return date_str

    m = re.match(r'^(\d{4}-\d{2}-\d{2})[T ]\d{2}:\d{2}:\d{2}', date_str)
    if m:
        return m.group(1)

    m = re.match(r'^(\d{1,2})/(\d{1,2})/(\d{4})$', date_str)
    if m:
        dd, mm, yyyy = m.groups()
        return f"{yyyy}-{mm.zfill(2)}-{dd.zfill(2)}"

    m = re.match(r'^(\d{4})/(\d{1,2})/(\d{1,2})$', date_str)
    if m:
        yyyy, mm, dd = m.groups()
        return f"{yyyy}-{mm.zfill(2)}-{dd.zfill(2)}"

    return date_str


def _write_temp_csv(rows_out, source_label: str) -> str:
    tmp = tempfile.NamedTemporaryFile(
        mode="w", suffix=".csv", delete=False, encoding="utf-8"
    )
    tmp.write("SqlTextInfo,Metric_Date,users\n")
    for rec in rows_out:
        tmp.write(rec + "\n")
    tmp.close()

    atexit.register(os.remove, tmp.name)
    print(f"[INFO] {source_label} detected — converted {len(rows_out)} rows -> {tmp.name}")
    return tmp.name


def _normalize_sql_field(sql_raw: str) -> str:
    """Guarantee exactly one trailing ';' and escape internal double-quotes
    so the field is a valid RFC-4180 quoted string."""
    sql_norm = sql_raw.strip().rstrip()
    if not sql_norm.endswith(";"):
        sql_norm += ";"
    return sql_norm.replace('"', '""')


def _convert_new_csv_to_input_format(src_path: str) -> str:
    rows_out = []
    with open(src_path, "r", encoding="utf-8", newline="") as fh:
        reader = csv.DictReader(fh)
        for row in reader:
            sql_raw  = (row.get("SqlTextInfo") or "").strip()
            log_date = _normalize_date((row.get("LogDate") or "").strip())
            user     = (row.get("user_name") or "").strip()

            if not sql_raw or not log_date or not user:
                continue

            sql_esc = _normalize_sql_field(sql_raw)
            rows_out.append(f'"{sql_esc}","{log_date}",{user}')

    return _write_temp_csv(rows_out, "New CSV schema")


def _convert_sql_users_csv_to_input_format(src_path: str) -> str:
    """
    Handles SqlTextInfo,Metric_Date,users files regardless of quoting —
    fully quoted (standard), fully unquoted, or a MIX of both across rows.

    csv.reader() respects RFC-4180 quoting per field/row (including
    embedded newlines inside quoted multiline SQL), so each row is read
    correctly whether or not that particular row happened to need quotes.

    The only case csv.reader can't disambiguate on its own is a row whose
    SQL contains a comma that was NOT quoted/escaped at all — that row
    parses into MORE than 3 fields. For those, we recombine: the last two
    fields are always Metric_Date and users (neither contains commas), so
    everything else — however many stray commas it has — is rejoined back
    into the SQL text.
    """
    rows_out = []
    with open(src_path, "r", encoding="utf-8", newline="") as fh:
        reader = csv.reader(fh)
        header_skipped = False
        for fields in reader:
            if not fields or not any((f or "").strip() for f in fields):
                continue  # skip blank lines

            if not header_skipped:
                header_skipped = True
                continue  # skip header row

            if len(fields) == 3:
                sql_raw, log_date, user = fields
            elif len(fields) > 3:
                # Unescaped comma(s) inside an unquoted SQL field — rejoin.
                sql_raw  = ",".join(fields[:-2])
                log_date = fields[-2]
                user     = fields[-1]
            else:
                continue  # malformed row — not enough fields, skip

            sql_raw  = sql_raw.strip()
            log_date = _normalize_date(log_date.strip())
            user     = user.strip()

            if not sql_raw or not log_date or not user:
                continue

            sql_esc = _normalize_sql_field(sql_raw)
            rows_out.append(f'"{sql_esc}","{log_date}",{user}')

    return _write_temp_csv(rows_out, "SqlTextInfo/Metric_Date/users CSV")


_fmt = _detect_format(INPUT_CSV)

if _fmt == "new_format":
    INPUT_CSV = _convert_new_csv_to_input_format(INPUT_CSV)
else:
    INPUT_CSV = _convert_sql_users_csv_to_input_format(INPUT_CSV)


# =============================================================================
# STEP 1 — Read & parse the multiline CSV
# =============================================================================

with open(INPUT_CSV, "r", encoding="utf-8") as fh:
    raw_text = fh.read()

body = "\n".join(raw_text.splitlines()[1:])     # drop header line

RECORD_RE = re.compile(
    r'"(.*?);"\s*,\s*"(\d{4}-\d{2}-\d{2})"\s*,\s*([^\s,"\n]+)\s*(?:\n|$)',
    re.DOTALL
)

raw_records = RECORD_RE.findall(body)
print(f"[INFO] Records found in CSV : {len(raw_records)}")


# =============================================================================
# STEP 2 — Helpers
# =============================================================================

def classify(username: str) -> str:
    return "app" if username.lower().startswith(APP_PREFIXES) else "user"


def _strip_locking_prefix(text: str) -> str:
    """
    Teradata locking-request modifiers (e.g. LOCKING ROW FOR ACCESS) are
    valid Teradata syntax at the TOP level of a statement, but sqlglot can't
    parse them when they appear nested inside a CREATE ... AS (...) body —
    that nesting is what causes the whole CREATE to fall back to a generic,
    unparseable "Command" node. Since the clause is just a lock hint (it
    doesn't affect which tables/columns are referenced), we strip it before
    re-parsing the inner query on its own.
    """
    stripped = text.lstrip()
    if re.match(r'(?i)^locking\b', stripped):
        m = re.search(r'(?i)\b(select|with)\b', stripped)
        if m:
            return stripped[m.start():]
    return text


def _extract_inner_query(sql_text: str):
    """
    Pull the SELECT/WITH body out of a CREATE ... AS (...) / CREATE ... AS
    SELECT ... statement so it can be parsed on its own, independent of the
    (possibly-unsupported-by-sqlglot) outer CREATE syntax.
    """
    m = re.search(r'\bAS\b\s*\(', sql_text, flags=re.IGNORECASE)
    if m:
        start = m.end() - 1  # index of '('
        depth = 0
        for i in range(start, len(sql_text)):
            if sql_text[i] == '(':
                depth += 1
            elif sql_text[i] == ')':
                depth -= 1
                if depth == 0:
                    return _strip_locking_prefix(sql_text[start + 1:i].strip())
        return None  # unbalanced parens — give up

    m = re.search(r'\bAS\b\s*(SELECT\b.*)', sql_text, flags=re.IGNORECASE | re.DOTALL)
    if m:
        return _strip_locking_prefix(m.group(1).strip())
    return None


def _extract_from_node(node) -> list:
    """Pull (table, column) pairs out of a parsed sqlglot node/subtree."""
    tables = [t.name.upper() for t in node.find_all(exp.Table) if t.name]
    cols   = list(dict.fromkeys(c.name.upper() for c in node.find_all(exp.Column) if c.name))
    stars  = list(node.find_all(exp.Star))

    pairs = []
    if stars and not cols:
        for tbl in tables:
            pairs.append((tbl, "*"))
    else:
        for tbl in tables:
            for col in cols:
                pairs.append((tbl, col))
    return pairs


def extract_table_column_pairs(raw_sql: str) -> tuple:
    """
    Returns (pairs, reason):
      pairs  — list of (TABLE, COLUMN) tuples, uppercased
      reason — None on success, string description if the SQL was unparseable

    Cases handled:
      CASE 1 — Genuine Teradata utility commands (SHOW/HELP/COLLECT/EXEC):
               sqlglot returns a Command node with no AST children.
               Logged & skipped, same as before.
      CASE 1b — CREATE ... AS (...) wrapping a LOCKING ROW FOR ACCESS (or
               similar) clause: sqlglot also returns Command for the WHOLE
               statement here, even though the inner query is perfectly
               valid Teradata SQL on its own. We recover by extracting the
               inner SELECT/WITH body, stripping the locking clause, and
               re-parsing just that — using ONLY the Teradata dialect (no
               cross-dialect fallback, to avoid other dialects silently
               "recovering" genuinely malformed SQL as something benign).
      CASE 2 — SELECT *: sqlglot returns Star nodes, not Column nodes →
               recorded as (TABLE, '*').
      CASE 3 — Normal SELECT/INSERT/UPDATE, and CREATE ... AS (SELECT ...)
               that parses fine directly → standard extraction, restricted
               to the query subtree (search_root) so the object being
               CREATEd isn't itself counted as a "referenced" table.
    """
    sql    = raw_sql.replace('""', '"').strip()
    sql    = re.sub(r'--[^\n]*', '', sql)       # strip  -- comments
    pairs  = []
    reason = None

    try:
        statements = sqlglot.parse(
            sql,
            dialect     = "teradata",
            error_level = sqlglot.ErrorLevel.IGNORE
        )
        for stmt in statements:
            if stmt is None:
                continue

            # CASE 1 / 1b: Teradata Command fallback
            if type(stmt).__name__ == "Command":
                inner_sql = _extract_inner_query(sql)
                recovered_pairs = []
                if inner_sql:
                    try:
                        inner_statements = sqlglot.parse(
                            inner_sql,
                            dialect     = "teradata",
                            error_level = sqlglot.ErrorLevel.IGNORE
                        )
                    except Exception:
                        inner_statements = []
                    if inner_statements and not any(
                        type(s).__name__ == "Command" for s in inner_statements if s is not None
                    ):
                        for inner_stmt in inner_statements:
                            if inner_stmt is not None:
                                recovered_pairs.extend(_extract_from_node(inner_stmt))

                if recovered_pairs:
                    pairs.extend(recovered_pairs)   # recovered — not a real skip
                else:
                    reason = f"Teradata utility command: {sql.strip()[:80]}"
                continue

            # search_root: for CREATE ... AS (SELECT ...), only look inside
            # the query part so the view/table name being CREATEd isn't
            # itself counted as a "referenced" table.
            search_root = stmt
            if isinstance(stmt, exp.Create):
                inner_expr = stmt.args.get("expression")
                if inner_expr is not None:
                    search_root = inner_expr

            pairs.extend(_extract_from_node(search_root))

    except Exception as exc:
        reason = f"Parse error: {exc}"

    if pairs:
        reason = None

    return pairs, reason


# =============================================================================
# STEP 3 — Explode every record → flat rows
# =============================================================================

exploded = []
skip_log = []

for raw_sql, date, user in raw_records:
    date = date.strip()
    user = user.strip()
    acct = classify(user)

    pairs, reason = extract_table_column_pairs(raw_sql)

    if reason:
        skip_log.append({"date": date, "user": user, "reason": reason,
                         "sql": raw_sql.replace('""', '"').strip()[:300]})

    for tbl, col in pairs:
        exploded.append({
            "Log_Date"   : date,
            "Table_Name" : tbl,
            "Column_Name": col,
            "username"   : user,
            "acct_type"  : acct,
        })

print(f"[INFO] Exploded rows (before agg) : {len(exploded)}")
print(f"[INFO] Skipped / unparseable      : {len(skip_log)}")

if not exploded:
    raise SystemExit("[ERROR] No rows after parsing — check INPUT_CSV path and format.")

df_exp = pd.DataFrame(exploded)


# =============================================================================
# STEP 4 — Aggregate by (Log_Date, Table_Name, Column_Name)
# =============================================================================

GRP_KEYS = ["Log_Date", "Table_Name", "Column_Name"]

df_users = (df_exp[df_exp["acct_type"] == "user"]
            .groupby(GRP_KEYS)["username"]
            .nunique()
            .rename("Distinct_Users"))

df_apps  = (df_exp[df_exp["acct_type"] == "app"]
            .groupby(GRP_KEYS)["username"]
            .nunique()
            .rename("Distinct_Apps"))

df_count = (df_exp.groupby(GRP_KEYS)["username"]
            .count()
            .rename("Usage_Count"))

df_agg = (pd.concat([df_count, df_users, df_apps], axis=1)
            .fillna(0)
            .astype({"Distinct_Users": int, "Distinct_Apps": int})
            .reset_index()
            .sort_values(GRP_KEYS)
            .reset_index(drop=True))


# =============================================================================
# STEP 5 — Add Row_Wid and Created_Timestamp
# =============================================================================

df_agg.insert(0, "Row_Wid", range(1, len(df_agg) + 1))
df_agg["Created_Timestamp"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

df_final = df_agg[[
    "Row_Wid", "Log_Date", "Table_Name", "Column_Name",
    "Usage_Count", "Distinct_Users", "Distinct_Apps", "Created_Timestamp",
]]

print(f"[INFO] Final aggregated rows      : {len(df_final)}")
print(df_final.head(10).to_string(index=False))


# =============================================================================
# STEP 6 — Write skip log
# =============================================================================

if skip_log:
    with open(SKIP_LOG, "w", encoding="utf-8") as f:
        f.write(f"Skipped queries — {datetime.now()}\n{'='*80}\n\n")
        for i, entry in enumerate(skip_log, 1):
            f.write(f"[{i}] date={entry['date']}  user={entry['user']}\n")
            f.write(f"     reason : {entry['reason']}\n")
            f.write(f"     sql    : {entry['sql']}\n\n")
    print(f"[INFO] Skip log written → {SKIP_LOG}")


# =============================================================================
# STEP 7 — Write to Excel
# =============================================================================

df_final.to_excel(OUTPUT_XLSX, index=False, sheet_name="Usage_Metrics")

wb = load_workbook(OUTPUT_XLSX)
ws = wb["Usage_Metrics"]

HEADER_FILL  = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT  = Font(name="Arial", bold=True, color="FFFFFF", size=10)
DATA_FONT    = Font(name="Arial", size=10)
ALT_FILL     = PatternFill("solid", fgColor="EBF3FB")
THIN_BORDER  = Border(
    left   = Side(style="thin", color="D9D9D9"),
    right  = Side(style="thin", color="D9D9D9"),
    top    = Side(style="thin", color="D9D9D9"),
    bottom = Side(style="thin", color="D9D9D9"),
)
NUMERIC_COLS = {1, 5, 6, 7}

for cell in ws[1]:
    cell.font      = HEADER_FONT
    cell.fill      = HEADER_FILL
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border    = THIN_BORDER
ws.row_dimensions[1].height = 22

for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
    fill = ALT_FILL if row_idx % 2 == 0 else PatternFill()
    for cell in row:
        cell.font      = DATA_FONT
        cell.fill      = fill
        cell.border    = THIN_BORDER
        cell.alignment = Alignment(
            horizontal = "center" if cell.column in NUMERIC_COLS else "left",
            vertical   = "center"
        )

for col_num, width in {1:10, 2:14, 3:42, 4:36, 5:14, 6:16, 7:14, 8:22}.items():
    ws.column_dimensions[get_column_letter(col_num)].width = width

ws.freeze_panes    = "A2"
ws.auto_filter.ref = ws.dimensions

wb.save(OUTPUT_XLSX)
print(f"[INFO] Saved → {OUTPUT_XLSX}")
