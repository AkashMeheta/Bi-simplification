# =============================================================================
# Teradata Query Usage Metrics Pipeline — Pure Python (No PySpark)
# =============================================================================
# Dependencies:  pip install pandas sqlglot openpyxl
#
# Output schema:
#   Row_Wid           – incremental integer (1, 2, 3 …)
#   Log_Date          – date of the query  (Metric_Date from CSV)
#   Table_Name        – table referenced in the SQL
#   Column_Name       – column referenced in the SQL  (* = SELECT * query)
#   Usage_Count       – # times this (table, col) pair appears on that date
#   Distinct_Users    – # unique human users on that date / table / column
#   Distinct_Apps     – # unique app accounts on that date / table / column
#   Created_Timestamp – timestamp when this script ran
#
# App detection rule (case-insensitive prefix match):
#   username starts with  svp | ovt | dt  →  APP,  else  →  USER
#
# Unparseable query types (logged to skip_log.txt, NOT dropped silently):
#   • SHOW TABLE / HELP TABLE / COLLECT STATISTICS / EXEC  (Teradata utilities)
#   • Completely malformed SQL that sqlglot cannot recover
# =============================================================================

import re
from datetime import datetime

import pandas as pd
import sqlglot
from sqlglot import exp
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── CONFIG ────────────────────────────────────────────────────────────────────

INPUT_CSV    = "input_3_c.csv"                  # ← your raw CSV from Teradata
OUTPUT_XLSX  = "teradata_usage_metrics.xlsx"    # ← output Excel
SKIP_LOG     = "skipped_queries.txt"            # ← unparseable queries logged here

APP_PREFIXES = ("svp", "ovt", "dt")            # case-insensitive prefix → app

# ─────────────────────────────────────────────────────────────────────────────
# STEP 1 — Read & parse the multiline CSV
# Format per record:  "<multiline SQL>;","YYYY-MM-DD",USERID
# ─────────────────────────────────────────────────────────────────────────────

with open(INPUT_CSV, "r", encoding="utf-8") as fh:
    raw_text = fh.read()

body = "\n".join(raw_text.splitlines()[1:])     # drop header line

RECORD_RE = re.compile(
    r'"(.*?);"\s*,\s*"([^"]+)"\s*,\s*(\S+)',
    re.DOTALL
)
raw_records = RECORD_RE.findall(body)           # → list of (sql_body, date, user)
print(f"[INFO] Records found in CSV : {len(raw_records)}")


# ─────────────────────────────────────────────────────────────────────────────
# STEP 2 — Helpers
# ─────────────────────────────────────────────────────────────────────────────

def classify(username: str) -> str:
    """'app' if username starts with a known app prefix, else 'user'."""
    return "app" if username.lower().startswith(APP_PREFIXES) else "user"


def extract_table_column_pairs(raw_sql: str) -> tuple[list, str | None]:
    """
    Parse one SQL string with sqlglot (Teradata dialect) and return:
      pairs  – list of (TABLE_NAME, COLUMN_NAME) tuples, uppercased
      reason – None if successful, or a string describing why it was skipped

    Handles 3 cases that previously caused silent "no pairs" drops:

    CASE 1 – Teradata utility statements (SHOW / HELP / COLLECT / EXEC):
      sqlglot returns a generic Command node with no AST children.
      These are genuine metadata queries with no table/column to extract.
      → pairs=[], reason="Teradata utility command …"

    CASE 2 – SELECT * with no explicit column list:
      sqlglot returns Star nodes instead of Column nodes, so cols=[].
      We record the table with a '*' wildcard so the row isn't dropped.
      → pairs=[(TABLE, '*')]

    CASE 3 – Normal queries with explicit columns:
      Standard cross-join of every table × every column found in the AST.
      → pairs=[(TABLE, COL), …]
    """
    sql = raw_sql.replace('""', '"').strip()
    sql = re.sub(r'--[^\n]*', '', sql)          # strip  -- line comments

    pairs  = []
    reason = None

    try:
        statements = sqlglot.parse(
            sql,
            dialect="teradata",                 # Teradata dialect: handles TOP, QUALIFY, etc.
            error_level=sqlglot.ErrorLevel.IGNORE
        )

        for stmt in statements:
            if stmt is None:
                continue

            # CASE 1: utility/DDL command — sqlglot can't recover table info
            if type(stmt).__name__ == "Command":
                reason = f"Teradata utility command: {sql.strip()[:80]}"
                continue

            tables = [t.name.upper() for t in stmt.find_all(exp.Table) if t.name]
            cols   = list(dict.fromkeys(            # dedupe, preserve order
                c.name.upper()
                for c in stmt.find_all(exp.Column)
                if c.name
            ))
            stars = list(stmt.find_all(exp.Star))

            # CASE 2: SELECT * — no explicit column nodes, only Star
            if stars and not cols:
                for tbl in tables:
                    pairs.append((tbl, "*"))        # '*' = full table scan / no specific col
                continue

            # CASE 3: normal query
            for tbl in tables:
                for col in cols:
                    pairs.append((tbl, col))

    except Exception as exc:
        reason = f"Parse error: {exc}"

    return pairs, reason


# ─────────────────────────────────────────────────────────────────────────────
# STEP 3 — Explode every record → (date, table, column, user, acct_type)
# ─────────────────────────────────────────────────────────────────────────────

exploded  = []
skip_log  = []      # collect all unparseable records for the log file

for raw_sql, date, user in raw_records:
    date = date.strip()
    user = user.strip()
    acct = classify(user)

    pairs, reason = extract_table_column_pairs(raw_sql)

    if reason:
        # Log it but don't crash — utility commands are expected
        skip_log.append({
            "date"  : date,
            "user"  : user,
            "reason": reason,
            "sql"   : raw_sql.replace('""', '"').strip()[:300],
        })

    if not pairs:
        continue    # nothing to aggregate for this record

    for tbl, col in pairs:
        exploded.append({
            "Log_Date"   : date,
            "Table_Name" : tbl,
            "Column_Name": col,
            "username"   : user,
            "acct_type"  : acct,       # "user" | "app"
        })

print(f"[INFO] Exploded rows (before agg) : {len(exploded)}")
print(f"[INFO] Skipped / unparseable      : {len(skip_log)}")


# ─────────────────────────────────────────────────────────────────────────────
# STEP 4 — Aggregate by (Log_Date, Table_Name, Column_Name)
# ─────────────────────────────────────────────────────────────────────────────

if not exploded:
    raise SystemExit("[ERROR] No rows after parsing — check INPUT_CSV path and format.")

df_exp = pd.DataFrame(exploded)

grp = df_exp.groupby(["Log_Date", "Table_Name", "Column_Name"], sort=True)

df_agg = grp.agg(
    Usage_Count    = ("username", "count"),
    Distinct_Users = ("username", lambda s:
        df_exp.loc[s.index].loc[df_exp.loc[s.index, "acct_type"] == "user", "username"].nunique()
    ),
    Distinct_Apps  = ("username", lambda s:
        df_exp.loc[s.index].loc[df_exp.loc[s.index, "acct_type"] == "app",  "username"].nunique()
    ),
).reset_index()


# ─────────────────────────────────────────────────────────────────────────────
# STEP 5 — Add Row_Wid and Created_Timestamp
# ─────────────────────────────────────────────────────────────────────────────

df_agg.insert(0, "Row_Wid", range(1, len(df_agg) + 1))
df_agg["Created_Timestamp"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

df_final = df_agg[[
    "Row_Wid", "Log_Date", "Table_Name", "Column_Name",
    "Usage_Count", "Distinct_Users", "Distinct_Apps", "Created_Timestamp",
]]

print(f"[INFO] Final aggregated rows      : {len(df_final)}")


# ─────────────────────────────────────────────────────────────────────────────
# STEP 6 — Write skip log (so you know exactly which queries were unresolvable)
# ─────────────────────────────────────────────────────────────────────────────

if skip_log:
    with open(SKIP_LOG, "w", encoding="utf-8") as f:
        f.write(f"Skipped queries — {datetime.now()}\n")
        f.write("=" * 80 + "\n\n")
        for i, entry in enumerate(skip_log, 1):
            f.write(f"[{i}] date={entry['date']}  user={entry['user']}\n")
            f.write(f"     reason : {entry['reason']}\n")
            f.write(f"     sql    : {entry['sql']}\n\n")
    print(f"[INFO] Skip log written → {SKIP_LOG}")


# ─────────────────────────────────────────────────────────────────────────────
# STEP 7 — Write to Excel with clean formatting
# ─────────────────────────────────────────────────────────────────────────────

df_final.to_excel(OUTPUT_XLSX, index=False, sheet_name="Usage_Metrics")

wb = load_workbook(OUTPUT_XLSX)
ws = wb["Usage_Metrics"]

HEADER_FILL  = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT  = Font(name="Arial", bold=True, color="FFFFFF", size=10)
DATA_FONT    = Font(name="Arial", size=10)
ALT_FILL     = PatternFill("solid", fgColor="EBF3FB")
THIN_BORDER  = Border(
    left   = Side(style="thin", color="D9D9D9"),
    right  = Side(style="thin", color="D9D9D9"),
    top    = Side(style="thin", color="D9D9D9"),
    bottom = Side(style="thin", color="D9D9D9"),
)
NUMERIC_COLS = {1, 5, 6, 7}    # Row_Wid, Usage_Count, Distinct_Users, Distinct_Apps

for cell in ws[1]:
    cell.font      = HEADER_FONT
    cell.fill      = HEADER_FILL
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border    = THIN_BORDER
ws.row_dimensions[1].height = 22

for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
    fill = ALT_FILL if row_idx % 2 == 0 else PatternFill()
    for cell in row:
        cell.font      = DATA_FONT
        cell.fill      = fill
        cell.border    = THIN_BORDER
        cell.alignment = Alignment(
            horizontal = "center" if cell.column in NUMERIC_COLS else "left",
            vertical   = "center"
        )

for col_num, width in {1:10, 2:14, 3:42, 4:36, 5:14, 6:16, 7:14, 8:22}.items():
    ws.column_dimensions[get_column_letter(col_num)].width = width

ws.freeze_panes    = "A2"
ws.auto_filter.ref = ws.dimensions

wb.save(OUTPUT_XLSX)
print(f"[INFO] Saved → {OUTPUT_XLSX}")


------------------------------------------------
import pandas as pd
import re

INPUT_FILE  = "input_3_c.csv"
OUTPUT_FILE = "output_parsed.xlsx"

# ── Step 1: Read the entire file as raw text ─────────────────────────────────
with open(INPUT_FILE, "r", encoding="utf-8") as f:
    content = f.read()

# ── Step 2: Skip the header line ─────────────────────────────────────────────
lines = content.splitlines()
# First line is:  SqlTextInfo,Metric_Date,users  → skip it
body = "\n".join(lines[1:])

# ── Step 3: Split into individual records ────────────────────────────────────
# From the CSV every record looks like:
#
#   "SELECT ...
#    ...
#    AND something = 'ESI';","2026-03-30",C8V4SJ
#
# i.e.  opening "  ...multiline SQL...  ;" , "date" , user
#
# Regex breakdown:
#   "           →  literal opening double-quote of the SQL field
#   (.*?)       →  SQL content (non-greedy, DOTALL so . matches newlines)
#   ;"          →  SQL always ends with semicolon then closing double-quote
#   \s*,\s*     →  comma separator (allow spaces)
#   "([^"]+)"   →  date field in double quotes
#   \s*,\s*     →  comma separator
#   (\S+)       →  user id (no spaces)

RECORD_PATTERN = re.compile(
    r'"(.*?);"\s*,\s*"([^"]+)"\s*,\s*(\S+)',
    re.DOTALL
)

records = RECORD_PATTERN.findall(body)
print(f"[INFO] Total records found: {len(records)}")

# ── Step 4: Clean each SQL block and build rows ───────────────────────────────
parsed_rows = []
skipped     = []

for i, (raw_sql, date, user) in enumerate(records):
    try:
        # raw_sql is everything between the outer CSV quotes (semicolon included at end)
        # It may contain "" (CSV-escaped double quotes) → unescape to single "
        sql = raw_sql.replace('""', '"')

        # Strip leading/trailing whitespace
        sql = sql.strip()

        # Collapse runs of 3+ blank lines to keep SQL readable
        sql = re.sub(r'\n{3,}', '\n\n', sql)

        date = date.strip()
        user = user.strip()

        if not sql:
            raise ValueError("Empty SQL after cleaning")

        parsed_rows.append({
            "SqlTextInfo": sql,
            "Metric_Date": date,
            "users":       user,
        })

    except Exception as e:
        skipped.append((i, str(e), raw_sql[:80]))

print(f"[INFO] Parsed: {len(parsed_rows)}  |  Skipped: {len(skipped)}")
if skipped:
    print("[WARN] Skipped records:")
    for idx, reason, preview in skipped:
        print(f"  #{idx}: {reason} | {preview!r}")

# ── Step 5: Build DataFrame ───────────────────────────────────────────────────
if not parsed_rows:
    raise SystemExit("[ERROR] No records parsed. Check INPUT_FILE path and record format.")

df = pd.DataFrame(parsed_rows, columns=["SqlTextInfo", "Metric_Date", "users"])

print(f"\n[INFO] DataFrame shape: {df.shape}")
print(df[["Metric_Date", "users"]].to_string())
print("\n--- First SQL preview (first 300 chars) ---")
print(df["SqlTextInfo"].iloc[0][:300])

# ── Step 6: Save to Excel ─────────────────────────────────────────────────────
df.to_excel(OUTPUT_FILE, index=False)
print(f"\n[INFO] Saved to {OUTPUT_FILE}")
