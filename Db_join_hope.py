# =============================================================================
# Teradata Query JOIN Metrics Pipeline — Pure Python (No PySpark)
# =============================================================================
# Dependencies:  pip install pandas sqlglot openpyxl
#
# Output schema:
#   Row_Wid                  – incremental integer (1, 2, 3 …)
#   Log_Date                 – date of the query (Metric_Date from CSV)
#   Left_Table_Database      – schema/db of the FROM (left) table
#   Left_Table_Name          – name of the FROM (left) table
#   Right_Table_Database     – schema/db of the JOIN (right) table
#   Right_Table_Name         – name of the JOIN (right) table
#   Right_Table_Columns      – comma-separated list of columns fetched/used
#                              from the right (JOIN) table anywhere in the
#                              query (select list, ON, WHERE, GROUP BY, etc.)
#   Join_Type                – INNER / LEFT / RIGHT / FULL / JOIN
#   Join_Keys                – comma-separated ON clause column names
#   Join_Count                – # times this exact join pair appears on that date
#   Distinct_Users            – # unique human users who ran a query with this join
#   Distinct_Apps             – # unique app accounts who ran a query with this join
#   Created_Timestamp         – timestamp when this script ran
#
# App detection rule (case-insensitive prefix match) — extend APP_PREFIXES below:
#   username starts with any prefix in APP_PREFIXES  →  APP,  else  →  USER
#
# Handles:
#   • Multiline SQL records in the CSV
#   • Simple JOINs, multiple JOINs, subquery JOINs, UNION queries
#   • Multi-key ON conditions  (a.id=b.id AND a.type=b.type → "ID, TYPE")
#   • Teradata-specific syntax (TOP, CAST FORMAT, QUALIFY …)
#   • Utility commands (SHOW/HELP/COLLECT) → skipped and logged
# =============================================================================

import re
from datetime import datetime

import pandas as pd
import sqlglot
from sqlglot import exp
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── CONFIG ────────────────────────────────────────────────────────────────────

INPUT_CSV   = "input_3_c.csv"                   # ← same CSV as column-usage script
OUTPUT_XLSX = "teradata_join_metrics.xlsx"
SKIP_LOG    = "join_skipped_queries.txt"

# Add any app/service account prefixes here (case-insensitive)
APP_PREFIXES = (
    "svp", "ovt", "dt",
    "svt",                  # SVT_ETL_SIARX_PRD style accounts
    "etl", "svc", "app",    # common service account patterns
)


# =============================================================================
# STEP 0 — Auto-detect & convert alternate CSV formats to expected input format
# =============================================================================
# Supports multiple input schemas, normalized into what STEP 1 expects:
#
#       SqlTextInfo,Metric_Date,users
#       "<multiline SQL ending in ;>","YYYY-MM-DD",USERID
#
# Format A — "new_format":
#   Columns: user_name, Db_nm, Tbl_nm, SqlTextInfo, LogDate, StartTime,
#            LastResponseTime
#   Detection: header contains "user_name" AND "logdate"
#
# Format B/C — "sql_users_format":
#   Columns: SqlTextInfo, Metric_Date, users — handled regardless of quoting:
#   fully quoted (standard), fully unquoted (select * from t1,29/02/2026,user1),
#   or a MIX of both within the same file (a CSV writer using RFC-4180 minimal
#   quoting only wraps a field in quotes when it needs to, e.g. the SQL
#   contains a comma — so some rows may be quoted and others not). Every row
#   is inspected and normalized on its own via csv.reader, so mixed quoting
#   is handled correctly instead of guessing the whole file's format from
#   just the first row.
#
# Dates are normalized to YYYY-MM-DD (required by STEP 1's RECORD_RE),
# handling DD/MM/YYYY and ISO-8601 timestamps like 2026-06-25T14:48:39.466Z.
#
# csv.field_size_limit() is also raised, since the default (~131072 bytes)
# is too small for large multiline SQL text fields in big exports.
# =============================================================================

import csv
import os
import sys
import tempfile
import atexit

# Default csv field size limit (~131072 bytes) is too small for large
# multiline SQL text fields in big exports — raise it. sys.maxsize can raise
# OverflowError on platforms where the C `long` is 32-bit (e.g. some Windows
# builds), so step down until it's accepted.
_new_limit = sys.maxsize
while True:
    try:
        csv.field_size_limit(_new_limit)
        break
    except OverflowError:
        _new_limit = int(_new_limit / 10)


def _peek_lines(csv_path: str, n: int = 2):
    lines = []
    with open(csv_path, "r", encoding="utf-8") as fh:
        for line in fh:
            if line.strip():
                lines.append(line)
            if len(lines) >= n:
                break
    return lines


def _detect_format(csv_path: str) -> str:
    """
    Only two schemas need distinguishing up front:
      • "new_format"       — user_name/Db_nm/Tbl_nm/SqlTextInfo/LogDate/... columns
      • "sql_users_format" — SqlTextInfo,Metric_Date,users columns (quoted,
                              unquoted, or a MIX of both within the same file)

    We deliberately do NOT try to decide "already fully quoted" vs "needs
    conversion" here. A CSV writer using RFC-4180 minimal quoting only wraps
    a field in quotes when it actually needs to — so a real file can easily
    have some rows quoted and others not. Peeking at just the first data row
    to guess the whole file's format misclassifies that kind of mixed file
    and silently drops every row after the first mismatch. Instead,
    _convert_sql_users_csv_to_input_format() below inspects EVERY row on
    its own and normalizes it individually, so mixed quoting is a non-issue.
    """
    lines = _peek_lines(csv_path, 1)
    if not lines:
        return "sql_users_format"

    header_clean = lines[0].strip().lower().replace('"', '')

    if "user_name" in header_clean and "logdate" in header_clean:
        return "new_format"

    return "sql_users_format"


def _normalize_date(date_str: str) -> str:
    """
    Normalize a date string to YYYY-MM-DD, the format STEP 1's RECORD_RE
    requires. Handles:
      • YYYY-MM-DD                          -> unchanged
      • YYYY-MM-DDTHH:MM:SS(.ffffff)?Z?      -> date part only
                                                (2026-06-25T14:48:39.466Z -> 2026-06-25)
      • DD/MM/YYYY                           -> rearranged (29/02/2026 -> 2026-02-29)
      • YYYY/MM/DD                           -> rearranged
    Unrecognized shapes are returned unchanged.
    """
    date_str = date_str.strip()

    if re.match(r'^\d{4}-\d{2}-\d{2}$', date_str):
        return date_str

    m = re.match(r'^(\d{4}-\d{2}-\d{2})[T ]\d{2}:\d{2}:\d{2}', date_str)
    if m:
        return m.group(1)

    m = re.match(r'^(\d{1,2})/(\d{1,2})/(\d{4})$', date_str)
    if m:
        dd, mm, yyyy = m.groups()
        return f"{yyyy}-{mm.zfill(2)}-{dd.zfill(2)}"

    m = re.match(r'^(\d{4})/(\d{1,2})/(\d{1,2})$', date_str)
    if m:
        yyyy, mm, dd = m.groups()
        return f"{yyyy}-{mm.zfill(2)}-{dd.zfill(2)}"

    return date_str


def _write_temp_csv(rows_out, source_label: str) -> str:
    tmp = tempfile.NamedTemporaryFile(
        mode="w", suffix=".csv", delete=False, encoding="utf-8"
    )
    tmp.write("SqlTextInfo,Metric_Date,users\n")
    for rec in rows_out:
        tmp.write(rec + "\n")
    tmp.close()

    atexit.register(os.remove, tmp.name)
    print(f"[INFO] {source_label} detected — converted {len(rows_out)} rows -> {tmp.name}")
    return tmp.name


def _normalize_sql_field(sql_raw: str) -> str:
    """Guarantee exactly one trailing ';' and escape internal double-quotes
    so the field is a valid RFC-4180 quoted string."""
    sql_norm = sql_raw.strip().rstrip()
    if not sql_norm.endswith(";"):
        sql_norm += ";"
    return sql_norm.replace('"', '""')


def _convert_new_csv_to_input_format(src_path: str) -> str:
    rows_out = []
    with open(src_path, "r", encoding="utf-8", newline="") as fh:
        reader = csv.DictReader(fh)
        for row in reader:
            sql_raw  = (row.get("SqlTextInfo") or "").strip()
            log_date = _normalize_date((row.get("LogDate") or "").strip())
            user     = (row.get("user_name") or "").strip()

            if not sql_raw or not log_date or not user:
                continue

            sql_esc = _normalize_sql_field(sql_raw)
            rows_out.append(f'"{sql_esc}","{log_date}",{user}')

    return _write_temp_csv(rows_out, "New CSV schema")


def _convert_sql_users_csv_to_input_format(src_path: str) -> str:
    """
    Handles SqlTextInfo,Metric_Date,users files regardless of quoting —
    fully quoted (standard), fully unquoted, or a MIX of both across rows.

    csv.reader() respects RFC-4180 quoting per field/row (including
    embedded newlines inside quoted multiline SQL), so each row is read
    correctly whether or not that particular row happened to need quotes.

    The only case csv.reader can't disambiguate on its own is a row whose
    SQL contains a comma that was NOT quoted/escaped at all — that row
    parses into MORE than 3 fields. For those, we recombine: the last two
    fields are always Metric_Date and users (neither contains commas), so
    everything else — however many stray commas it has — is rejoined back
    into the SQL text.
    """
    rows_out = []
    with open(src_path, "r", encoding="utf-8", newline="") as fh:
        reader = csv.reader(fh)
        header_skipped = False
        for fields in reader:
            if not fields or not any((f or "").strip() for f in fields):
                continue  # skip blank lines

            if not header_skipped:
                header_skipped = True
                continue  # skip header row

            if len(fields) == 3:
                sql_raw, log_date, user = fields
            elif len(fields) > 3:
                # Unescaped comma(s) inside an unquoted SQL field — rejoin.
                sql_raw  = ",".join(fields[:-2])
                log_date = fields[-2]
                user     = fields[-1]
            else:
                continue  # malformed row — not enough fields, skip

            sql_raw  = sql_raw.strip()
            log_date = _normalize_date(log_date.strip())
            user     = user.strip()

            if not sql_raw or not log_date or not user:
                continue

            sql_esc = _normalize_sql_field(sql_raw)
            rows_out.append(f'"{sql_esc}","{log_date}",{user}')

    return _write_temp_csv(rows_out, "SqlTextInfo/Metric_Date/users CSV")


_fmt = _detect_format(INPUT_CSV)

if _fmt == "new_format":
    INPUT_CSV = _convert_new_csv_to_input_format(INPUT_CSV)
else:
    INPUT_CSV = _convert_sql_users_csv_to_input_format(INPUT_CSV)


# =============================================================================
# STEP 1 — Read & parse the multiline CSV
# =============================================================================
# Record format:
#   "SELECT ...
#    multiline SQL ...
#    WHERE x='Y';","YYYY-MM-DD",USERID
#
# Regex anchors:
#   • date field must match exactly YYYY-MM-DD
#   • user token must be followed only by whitespace/end-of-line
#   This prevents false splits on semicolons inside SQL string literals.
# =============================================================================

with open(INPUT_CSV, "r", encoding="utf-8") as fh:
    raw_text = fh.read()

body = "\n".join(raw_text.splitlines()[1:])     # drop header line

RECORD_RE = re.compile(
    r'"(.*?);"\s*,\s*"(\d{4}-\d{2}-\d{2})"\s*,\s*([^\s,"\n]+)\s*(?:\n|$)',
    re.DOTALL
)

raw_records = RECORD_RE.findall(body)
print(f"[INFO] Records found in CSV : {len(raw_records)}")


# =============================================================================
# STEP 2 — Helpers
# =============================================================================

def classify(username: str) -> str:
    """'app' if username starts with any known app prefix, else 'user'."""
    return "app" if username.lower().startswith(APP_PREFIXES) else "user"


def get_db_and_table(tbl_node) -> tuple:
    """Extract (DATABASE, TABLE_NAME) from a sqlglot Table AST node."""
    db   = tbl_node.args.get("db")
    name = tbl_node.name
    return (
        db.name.upper()   if db   else "",
        name.upper()      if name else "",
    )


def resolve_join_table(join_this):
    """
    Return the physical Table node for the right-hand side of a JOIN.

    • Direct table:  JOIN db.tbl t  → returns the Table node
    • Subquery:      JOIN (SELECT ... FROM db.tbl) t
                     → resolves to the subquery's own FROM table so we always
                        record real table names, never anonymous subqueries.
    """
    if isinstance(join_this, exp.Table):
        return join_this

    if isinstance(join_this, exp.Subquery):
        inner = join_this.this
        if isinstance(inner, exp.Select):
            from_node = inner.args.get("from_")
            if from_node and isinstance(from_node.this, exp.Table):
                return from_node.this

    return None


def extract_right_table_columns(select, alias: str) -> str:
    """
    Find every column in this SELECT scope (select list, WHERE, ON,
    GROUP BY, HAVING, QUALIFY, ORDER BY — find_all walks the whole
    subtree) that is qualified with the given table/alias, i.e. the
    columns actually "fetched"/referenced against the right-hand JOIN
    table.

    e.g.  SELECT a.id, b.name, b.amt FROM t1 a JOIN t2 b ON a.id=b.id
          alias="b"  →  "NAME, AMT, ID"  (any column referenced on b,
          including join keys, is included)

    Deduped, preserving first-seen order. Returns "" if no alias or
    no qualified columns are found (e.g. columns referenced without a
    table qualifier).
    """
    if not alias:
        return ""
    cols = []
    for col in select.find_all(exp.Column):
        tbl_node = col.args.get("table")
        tbl_name = tbl_node.name if tbl_node else None
        if tbl_name and tbl_name.upper() == alias.upper() and col.name:
            cols.append(col.name.upper())
    return ", ".join(dict.fromkeys(cols))


def extract_join_keys(on_expr) -> str:
    """
    Extract column names used as join keys from an ON expression.
    e.g.  a.id = b.id AND a.type = b.type  →  "ID, TYPE"
    Deduped, preserving first-seen order.
    """
    if on_expr is None:
        return ""
    keys = []
    for eq in on_expr.find_all(exp.EQ):
        for col in eq.find_all(exp.Column):
            if col.name:
                keys.append(col.name.upper())
    return ", ".join(dict.fromkeys(keys))


def extract_joins(raw_sql: str) -> tuple:
    """
    Parse one SQL string and return (join_list, skip_reason).

    join_list — list of dicts with keys:
        left_db, left_tbl, right_db, right_tbl, join_type, join_keys,
        right_columns

    skip_reason — None if OK, string if the SQL was unparseable/utility.

    Handles:
        • Simple JOINs, multiple JOINs per SELECT
        • Subquery JOINs  (resolves to subquery's own FROM table)
        • UNION queries   (walks every SELECT scope independently)
        • Multi-key ON    (a.id=b.id AND a.type=b.type)
        • Right-table column usage (any column referenced against the
          JOIN table's alias anywhere in that SELECT scope)
    """
    sql         = raw_sql.replace('""', '"').strip()
    sql         = re.sub(r'--[^\n]*', '', sql)      # strip line comments
    results     = []
    skip_reason = None

    try:
        statements = sqlglot.parse(
            sql,
            dialect     = "teradata",
            error_level = sqlglot.ErrorLevel.IGNORE,
        )

        for stmt in statements:
            if stmt is None:
                continue

            # Utility commands have no parseable structure
            if type(stmt).__name__ == "Command":
                skip_reason = f"Teradata utility command: {sql.strip()[:80]}"
                continue

            # Walk every SELECT scope — handles UNIONs and nested subqueries
            for select in stmt.find_all(exp.Select):
                from_node = select.args.get("from_")
                if not from_node or not isinstance(from_node.this, exp.Table):
                    continue

                left_db, left_tbl = get_db_and_table(from_node.this)
                if not left_tbl:
                    continue

                for join in select.args.get("joins") or []:
                    right_node = resolve_join_table(join.this)
                    if right_node is None:
                        continue

                    right_db, right_tbl = get_db_and_table(right_node)
                    if not right_tbl:
                        continue

                    join_type = str(join.args.get("kind") or "").upper() or "JOIN"
                    join_keys = extract_join_keys(join.args.get("on"))

                    # Alias actually used to qualify columns in the SQL
                    # (falls back to the real table name when no alias
                    # was assigned to the JOIN'd table/subquery).
                    right_alias = join.this.alias_or_name or right_tbl
                    right_columns = extract_right_table_columns(select, right_alias)

                    results.append({
                        "left_db"      : left_db,
                        "left_tbl"     : left_tbl,
                        "right_db"     : right_db,
                        "right_tbl"    : right_tbl,
                        "join_type"    : join_type,
                        "join_keys"    : join_keys,
                        "right_columns": right_columns,
                    })

    except Exception as exc:
        skip_reason = f"Parse error: {exc}"

    return results, skip_reason


# =============================================================================
# STEP 3 — Explode every record → flat join rows
# =============================================================================

exploded = []
skip_log = []

for raw_sql, date, user in raw_records:
    date = date.strip()
    user = user.strip()
    acct = classify(user)

    joins, skip_reason = extract_joins(raw_sql)

    if skip_reason:
        skip_log.append({
            "date"  : date,
            "user"  : user,
            "reason": skip_reason,
            "sql"   : raw_sql.replace('""', '"').strip()[:300],
        })

    if not joins:
        continue

    for j in joins:
        exploded.append({
            "Log_Date"     : date,
            "left_db"      : j["left_db"],
            "left_tbl"     : j["left_tbl"],
            "right_db"     : j["right_db"],
            "right_tbl"    : j["right_tbl"],
            "join_type"    : j["join_type"],
            "join_keys"    : j["join_keys"],
            "right_columns": j["right_columns"],
            "username"     : user,
            "acct_type"    : acct,
        })

print(f"[INFO] Exploded join rows (before agg) : {len(exploded)}")
print(f"[INFO] Skipped / unparseable           : {len(skip_log)}")

if not exploded:
    raise SystemExit("[ERROR] No join rows extracted. Check INPUT_CSV path and format.")

df_exp = pd.DataFrame(exploded)


# =============================================================================
# STEP 4 — Aggregate
# =============================================================================
# Group by (Log_Date, left_db, left_tbl, right_db, right_tbl, join_type, join_keys)
# Compute: Join_Count, Distinct_Users, Distinct_Apps, Right_Table_Columns
#
# Uses three separate groupby passes (avoids the lambda-closure bug where
# every group reports distinct count = 1), plus a fourth pass that unions
# every right-table column string seen within a group (different queries
# hitting the same join pair may reference different column subsets).
# =============================================================================

GRP_KEYS = [
    "Log_Date",
    "left_db", "left_tbl",
    "right_db", "right_tbl",
    "join_type", "join_keys",
]

df_users = (df_exp[df_exp["acct_type"] == "user"]
            .groupby(GRP_KEYS)["username"]
            .nunique()
            .rename("Distinct_Users"))

df_apps  = (df_exp[df_exp["acct_type"] == "app"]
            .groupby(GRP_KEYS)["username"]
            .nunique()
            .rename("Distinct_Apps"))

df_count = (df_exp.groupby(GRP_KEYS)["username"]
            .count()
            .rename("Join_Count"))


def _union_columns(series: pd.Series) -> str:
    """Union all comma-separated column lists within a group, deduped,
    preserving first-seen order across rows."""
    seen = {}
    for val in series:
        if not val:
            continue
        for col in val.split(","):
            col = col.strip()
            if col:
                seen[col] = True
    return ", ".join(seen.keys())


df_right_cols = (df_exp.groupby(GRP_KEYS)["right_columns"]
                  .apply(_union_columns)
                  .rename("Right_Table_Columns"))

df_agg = (pd.concat([df_count, df_users, df_apps, df_right_cols], axis=1)
            .fillna({"Distinct_Users": 0, "Distinct_Apps": 0, "Right_Table_Columns": ""})
            .astype({"Distinct_Users": int, "Distinct_Apps": int})
            .reset_index()
            .sort_values(GRP_KEYS)
            .reset_index(drop=True))


# =============================================================================
# STEP 5 — Rename columns to final schema & add Row_Wid / Created_Timestamp
# =============================================================================

df_agg.insert(0, "Row_Wid", range(1, len(df_agg) + 1))
df_agg["Created_Timestamp"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

df_final = df_agg.rename(columns={
    "left_db"  : "Left_Table_Database",
    "left_tbl" : "Left_Table_Name",
    "right_db" : "Right_Table_Database",
    "right_tbl": "Right_Table_Name",
    "join_type": "Join_Type",
    "join_keys": "Join_Keys",
})[[
    "Row_Wid",
    "Log_Date",
    "Left_Table_Database",
    "Left_Table_Name",
    "Right_Table_Database",
    "Right_Table_Name",
    "Right_Table_Columns",
    "Join_Type",
    "Join_Keys",
    "Join_Count",
    "Distinct_Users",
    "Distinct_Apps",
    "Created_Timestamp",
]]

print(f"[INFO] Final aggregated rows : {len(df_final)}")
print(df_final.head(10).to_string(index=False))


# =============================================================================
# STEP 6 — Write skip log
# =============================================================================

if skip_log:
    with open(SKIP_LOG, "w", encoding="utf-8") as f:
        f.write(f"Skipped queries — {datetime.now()}\n{'='*80}\n\n")
        for i, entry in enumerate(skip_log, 1):
            f.write(f"[{i}] date={entry['date']}  user={entry['user']}\n")
            f.write(f"     reason : {entry['reason']}\n")
            f.write(f"     sql    : {entry['sql']}\n\n")
    print(f"[INFO] Skip log written → {SKIP_LOG}")


# =============================================================================
# STEP 7 — Write to Excel with formatting
# =============================================================================

df_final.to_excel(OUTPUT_XLSX, index=False, sheet_name="Join_Metrics")

wb = load_workbook(OUTPUT_XLSX)
ws = wb["Join_Metrics"]

HEADER_FILL  = PatternFill("solid", fgColor="1F3864")   # deep navy
HEADER_FONT  = Font(name="Arial", bold=True, color="FFFFFF", size=10)
DATA_FONT    = Font(name="Arial", size=10)
ALT_FILL     = PatternFill("solid", fgColor="E9EFF7")
THIN_BORDER  = Border(
    left   = Side(style="thin", color="D0D0D0"),
    right  = Side(style="thin", color="D0D0D0"),
    top    = Side(style="thin", color="D0D0D0"),
    bottom = Side(style="thin", color="D0D0D0"),
)

# Col indices for centre-align (1-based):
# Row_Wid(1), Join_Type(8), Join_Count(10), Distinct_Users(11), Distinct_Apps(12)
CENTRE_COLS = {1, 8, 10, 11, 12}

# Header row
for cell in ws[1]:
    cell.font      = HEADER_FONT
    cell.fill      = HEADER_FILL
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border    = THIN_BORDER
ws.row_dimensions[1].height = 30

# Data rows
for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
    fill = ALT_FILL if row_idx % 2 == 0 else PatternFill()
    for cell in row:
        cell.font      = DATA_FONT
        cell.fill      = fill
        cell.border    = THIN_BORDER
        cell.alignment = Alignment(
            horizontal = "center" if cell.column in CENTRE_COLS else "left",
            vertical   = "center",
        )

# Column widths
col_widths = {
    1 : 9,   # Row_Wid
    2 : 13,  # Log_Date
    3 : 22,  # Left_Table_Database
    4 : 35,  # Left_Table_Name
    5 : 22,  # Right_Table_Database
    6 : 35,  # Right_Table_Name
    7 : 40,  # Right_Table_Columns
    8 : 10,  # Join_Type
    9 : 30,  # Join_Keys
    10: 12,  # Join_Count
    11: 15,  # Distinct_Users
    12: 14,  # Distinct_Apps
    13: 22,  # Created_Timestamp
}
for col_num, width in col_widths.items():
    ws.column_dimensions[get_column_letter(col_num)].width = width

ws.freeze_panes    = "A2"
ws.auto_filter.ref = ws.dimensions

wb.save(OUTPUT_XLSX)
print(f"[INFO] Saved → {OUTPUT_XLSX}")
