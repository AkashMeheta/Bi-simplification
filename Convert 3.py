"""
SQL Parsing & Column Usage Metrics Pipeline
============================================
Parses SQL queries from a DataFrame, extracts table/column usage,
and outputs structured metrics to Excel.
"""

import re
import logging
import pandas as pd
import sqlglot
import sqlglot.expressions as exp
from collections import defaultdict
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ---------------------------------------------------------------------------
# Logging
# ---------------------------------------------------------------------------
logging.basicConfig(
    filename="sql_parse_errors.log",
    level=logging.WARNING,
    format="%(asctime)s | %(levelname)s | %(message)s",
)
logger = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# Classification
# ---------------------------------------------------------------------------
def classify_user(username: str) -> str:
    """Return 'App' if username contains 'svt' or 'opt' (case-insensitive), else 'User'."""
    lower = (username or "").lower()
    return "App" if ("svt" in lower or "opt" in lower) else "User"


# ---------------------------------------------------------------------------
# SQL Pre-processing
# ---------------------------------------------------------------------------
_SINGLE_LINE_COMMENT = re.compile(r"--[^\n]*")
_MULTI_LINE_COMMENT = re.compile(r"/\*.*?\*/", re.DOTALL)


def strip_comments_safe(sql: str) -> str:
    """
    Remove SQL comments without collapsing line breaks.
    Single-line comments (--) are replaced with a newline so that the
    next token is NOT accidentally merged with the previous line.
    Multi-line comments are replaced with a space.
    """
    sql = _MULTI_LINE_COMMENT.sub(" ", sql)
    sql = _SINGLE_LINE_COMMENT.sub("\n", sql)
    return sql


# ---------------------------------------------------------------------------
# Core Parser
# ---------------------------------------------------------------------------
def extract_table_column_usage(sql: str) -> dict[tuple[str, str], int]:
    """
    Parse *one* SQL statement and return a dict of
        {(table_name_lower, column_name_lower): count}

    Handles:
    - Quoted identifiers
    - Aliases (column aliases are excluded; only real column refs counted)
    - Multi-line SQL
    - Comments (stripped safely before parsing)
    - Column references without explicit table qualifier

    Returns an empty dict on parse failure (logged separately).
    """
    cleaned = strip_comments_safe(sql).strip()
    if not cleaned:
        return {}

    try:
        statements = sqlglot.parse(cleaned, error_level=sqlglot.ErrorLevel.RAISE)
    except Exception as e:
        raise ValueError(f"sqlglot parse error: {e}") from e

    usage: dict[tuple[str, str], int] = defaultdict(int)

    for stmt in statements:
        if stmt is None:
            continue

        # Build alias → real table mapping from FROM / JOIN clauses
        alias_map: dict[str, str] = {}
        for table_expr in stmt.find_all(exp.Table):
            tname = (table_expr.name or "").lower().strip('"').strip("'")
            alias = (table_expr.alias or "").lower().strip('"').strip("'")
            if alias:
                alias_map[alias] = tname
            if tname:
                alias_map[tname] = tname  # self-mapping for unaliased tables

        # Walk every Column node
        for col_expr in stmt.find_all(exp.Column):
            col_name = (col_expr.name or "").lower().strip('"').strip("'")
            if not col_name or col_name == "*":
                continue

            # Resolve table qualifier
            table_qualifier = ""
            if col_expr.table:
                raw_tq = col_expr.table.lower().strip('"').strip("'")
                table_qualifier = alias_map.get(raw_tq, raw_tq)

            # If no qualifier, try to infer from single-table queries
            if not table_qualifier:
                real_tables = [v for v in alias_map.values() if v]
                if len(real_tables) == 1:
                    table_qualifier = real_tables[0]
                else:
                    table_qualifier = "_unknown_"

            usage[(table_qualifier, col_name)] += 1

    return dict(usage)


# ---------------------------------------------------------------------------
# Main Pipeline
# ---------------------------------------------------------------------------
def process_dataframe(df: pd.DataFrame) -> pd.DataFrame:
    """
    Process the input DataFrame and return the metrics DataFrame.

    Expected input columns: LogDate, UserName, SqlTextInfo
    Output columns: RowNumber, MetricDate, TableName, ColumnName,
                    UsageCount, UniqueUserCount, UniqueAppCount
    """
    # Accumulate: {(MetricDate, TableName, ColumnName) -> {users: set, apps: set, count: int}}
    aggregator: dict[tuple, dict] = defaultdict(
        lambda: {"count": 0, "users": set(), "apps": set()}
    )

    for row_num, row in df.iterrows():
        sql_text = str(row.get("SqlTextInfo", "") or "")
        username = str(row.get("UserName", "") or "")
        log_date = row.get("LogDate")

        # Normalise date to date-only
        if pd.isnull(log_date):
            metric_date = None
        elif hasattr(log_date, "date"):
            metric_date = log_date.date()
        else:
            try:
                metric_date = pd.to_datetime(log_date).date()
            except Exception:
                metric_date = None

        classification = classify_user(username)

        try:
            usage = extract_table_column_usage(sql_text)
        except Exception as e:
            logger.warning(
                "Row %s | User: %s | Date: %s | Error: %s | SQL snippet: %.200s",
                row_num,
                username,
                metric_date,
                e,
                sql_text,
            )
            usage = {}

        for (table_name, col_name), count in usage.items():
            key = (metric_date, table_name, col_name)
            aggregator[key]["count"] += count
            if classification == "User":
                aggregator[key]["users"].add(username)
            else:
                aggregator[key]["apps"].add(username)

    # Build output records
    records = []
    for row_num, ((metric_date, table_name, col_name), data) in enumerate(
        aggregator.items(), start=1
    ):
        records.append(
            {
                "RowNumber": row_num,
                "MetricDate": metric_date,
                "TableName": table_name,
                "ColumnName": col_name,
                "UsageCount": data["count"],
                "UniqueUserCount": len(data["users"]),
                "UniqueAppCount": len(data["apps"]),
            }
        )

    return pd.DataFrame(records)


# ---------------------------------------------------------------------------
# Excel Writer
# ---------------------------------------------------------------------------
_HEADER_FILL = PatternFill("solid", start_color="1F4E79")
_ALT_FILL = PatternFill("solid", start_color="D6E4F0")
_THIN = Side(style="thin", color="BFBFBF")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

COL_WIDTHS = {
    "RowNumber": 12,
    "MetricDate": 16,
    "TableName": 30,
    "ColumnName": 30,
    "UsageCount": 14,
    "UniqueUserCount": 18,
    "UniqueAppCount": 16,
}


def write_excel(metrics_df: pd.DataFrame, output_path: str) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "SQL Column Metrics"

    headers = list(COL_WIDTHS.keys())

    # --- Header row ---
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = COL_WIDTHS[header]

    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A2"

    # --- Data rows ---
    for r_idx, row in metrics_df.iterrows():
        excel_row = r_idx + 2  # offset for header
        fill = _ALT_FILL if (r_idx % 2 == 0) else PatternFill()

        for c_idx, col in enumerate(headers, start=1):
            val = row[col]
            # Convert date objects for Excel compatibility
            if hasattr(val, "strftime"):
                val = val.strftime("%Y-%m-%d")
            cell = ws.cell(row=excel_row, column=c_idx, value=val)
            cell.font = Font(name="Arial", size=10)
            cell.fill = fill
            cell.border = _BORDER
            cell.alignment = Alignment(
                horizontal="center" if col in ("RowNumber", "MetricDate", "UsageCount",
                                               "UniqueUserCount", "UniqueAppCount")
                else "left"
            )

    # --- Auto-filter ---
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    wb.save(output_path)
    print(f"✅ Excel written → {output_path}  ({len(metrics_df):,} rows)")


# ---------------------------------------------------------------------------
# Entry Point
# ---------------------------------------------------------------------------
def run_pipeline(df: pd.DataFrame, output_path: str = "sql_column_metrics.xlsx") -> pd.DataFrame:
    """
    End-to-end pipeline.
    df must have columns: LogDate, UserName, SqlTextInfo
    Returns the metrics DataFrame (also written to output_path).
    """
    print(f"Processing {len(df):,} input rows …")
    metrics_df = process_dataframe(df)
    print(f"Extracted {len(metrics_df):,} (date, table, column) metric records.")
    write_excel(metrics_df, output_path)
    return metrics_df


# ---------------------------------------------------------------------------
# Demo / Smoke-test
# ---------------------------------------------------------------------------
if __name__ == "__main__":
    sample_data = {
        "LogDate": [
            "2024-01-15", "2024-01-15", "2024-01-16",
            "2024-01-16", "2024-01-17", "2024-01-17",
        ],
        "UserName": [
            "john.doe",       # User
            "svt_service01",  # App  (contains 'svt')
            "jane.smith",     # User
            "opt_runner",     # App  (contains 'opt')
            "bob.jones",      # User
            "SVT_BATCH",      # App  (case-insensitive)
        ],
        "SqlTextInfo": [
            # Multi-line with double-quoted identifiers
            """
            SELECT
                o."order_id",
                o."customer_id",
                -- inline comment should not collapse next line
                o."total_amount"
            FROM orders o
            WHERE o."status" = 'active'
            """,

            # Multi-line with a block comment
            """
            SELECT
                c.customer_id,
                c.first_name,
                c.last_name,
                /* this is a
                   multi-line comment */
                c.email
            FROM customers c
            JOIN orders o ON c.customer_id = o.customer_id
            WHERE c.region = 'APAC'
            """,

            # Subquery
            """
            SELECT p.product_name, p.price, inv.quantity
            FROM products p
            LEFT JOIN inventory inv ON p.product_id = inv.product_id
            WHERE p.category = 'electronics'
              AND inv.quantity > 0
            """,

            # Malformed SQL – should be logged, not crash
            "SELECT * FROM WHERE broken $$$ syntax",

            # Aggregate with alias (alias columns should not count as column refs)
            """
            SELECT
                s.region,
                SUM(s.revenue)   AS total_revenue,
                COUNT(s.sale_id) AS num_sales
            FROM sales s
            GROUP BY s.region
            """,

            # Simple single-table – no explicit qualifier
            """
            SELECT employee_id, first_name, last_name, department
            FROM employees
            WHERE hire_date >= '2020-01-01'
            """,
        ],
    }

    df_input = pd.DataFrame(sample_data)
    result = run_pipeline(df_input, "sql_column_metrics.xlsx")
    print("\nSample output (first 20 rows):")
    print(result.head(20).to_string(index=False))
